#!/usr/bin/env python3
"""
Generates the Ostseeliebe Buchungs-Dashboard HTML from a CSV booking export.
No external dependencies required - uses only Python standard library.
"""

import csv
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Paths - resolve the mapped session paths to real filesystem paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Default paths - can be overridden via command-line args
DEFAULT_CSV = "buchungen_export_2027.csv"
DEFAULT_OUT = "index.html"

# ---------------------------------------------------------------------------
# Apartmenthaus → Unterkunft Mapping (Objekt-Nummern aus Contao)
# ---------------------------------------------------------------------------
APARTMENTHAUS_MAPPING = {
    "Apartmenthaus Viktoria Luna": [{"nr":174,"name":"Elli"},{"nr":140,"name":"Kleine Lagune"}],
    "Darßer Landhaus": [{"nr":269,"name":"Ruheoase"},{"nr":270,"name":"Wellnessoase"}],
    "Das Feriendomizil Traumzeit Zingst": [{"nr":218,"name":"Traumzeit 1"},{"nr":256,"name":"Traumzeit 2"}],
    "Das Feriendomizil Zingst – Strandhafer & Strandkaten": [{"nr":259,"name":"RC - Strandhafer"},{"nr":258,"name":"RC - Strandkaten"}],
    "Ferienhäuser Alter Schwede": [{"nr":201,"name":"Alter Schwede 1"},{"nr":202,"name":"Alter Schwede 2"},{"nr":203,"name":"Alter Schwede 3"}],
    "Ferienhäuser grüner Winkel": [{"nr":56,"name":"RC - Charlotte"},{"nr":57,"name":"RC - Therese"}],
    "Ferienwohnung Velo": [{"nr":77,"name":"Velo 1"},{"nr":78,"name":"Velo 2"},{"nr":79,"name":"Velo 3"},{"nr":138,"name":"Velo 4"}],
    "Haus am kleinen Hafen": [{"nr":273,"name":"kl. Hafen No 1"},{"nr":274,"name":"kl. Hafen No 2"},{"nr":275,"name":"kl. Hafen No 3"}],
    "Haus Bi de Wisch": [{"nr":221,"name":"Bi de Wisch EG"},{"nr":222,"name":"Bi de Wisch OG"}],
    "Haus Blaue Wieck": [{"nr":260,"name":"DWR Mitte"},{"nr":261,"name":"OSS rechts"},{"nr":262,"name":"WSL links"}],
    "Haus Bliesenrade": [{"nr":271,"name":"Bernstein"}],
    "Haus Boddenblick": [{"nr":283,"name":"Boddenblick 9"}],
    "Haus Chausseestraße": [{"nr":281,"name":"Moin Moin"},{"nr":278,"name":"Santa Karina Born"}],
    "Haus Citynah": [{"nr":137,"name":"Frau Zander"},{"nr":194,"name":"Möwennest Nr. 5"},{"nr":166,"name":"Seewind"}],
    "Haus Cozy Five": [{"nr":190,"name":"Cozy 1"},{"nr":191,"name":"Cozy 2"},{"nr":192,"name":"Cozy 3"},{"nr":131,"name":"Cozy 4"},{"nr":132,"name":"Cozy 5"}],
    "Haus Darssduett": [{"nr":279,"name":"Darss-Duett 1"},{"nr":290,"name":"Darss-Duett 2"}],
    "Haus Darßer Sonnenfisch": [{"nr":228,"name":"Clownfisch"},{"nr":247,"name":"Sonnendeck"}],
    "Häuser Haseneck": [{"nr":129,"name":"DAT KROEGER HUS"},{"nr":139,"name":"Kranschehus"}],
    "Haus Hoppenberg Strandquartier": [{"nr":145,"name":"Bärbel"},{"nr":144,"name":"Lotta"}],
    "Haus im Zentrum": [{"nr":114,"name":"Ocean Star"},{"nr":99,"name":"Rewal"},{"nr":115,"name":"Strandhaus Zingst"}],
    "Haus In den Wiesen": [{"nr":282,"name":"In den Wiesen App. 2"},{"nr":227,"name":"In den Wiesen 3"}],
    "Haus Kraanstiet": [{"nr":293,"name":"Kraanstiet 1"},{"nr":294,"name":"Kraanstiet 2"},{"nr":295,"name":"Kraanstiet 3"}],
    "Haus Küstentour": [{"nr":90,"name":"Dünenläufer"},{"nr":92,"name":"Ostseekoje"},{"nr":94,"name":"Sandbank No. 4"},{"nr":91,"name":"Strandsegler"}],
    "Haus Küstenzauber": [{"nr":98,"name":"Bremen"},{"nr":97,"name":"Stralsund"}],
    "Haus Meerkaten": [{"nr":168,"name":"Lust auf Meer"},{"nr":148,"name":"Schifferkaten"}],
    "Haus Meerle": [{"nr":288,"name":"RC - Meerle 1"},{"nr":291,"name":"RC - Meerle 2"},{"nr":296,"name":"RC - Meerle 3"},{"nr":297,"name":"RC - Meerle 4"},{"nr":299,"name":"RC - Meerle 5"}],
    "Haus Öresundhus": [{"nr":267,"name":"Öresundhus Whg.2"},{"nr":268,"name":"Öresundhus Whg.3"}],
    "Haus Quartett Küstenglück": [{"nr":195,"name":"Windland"},{"nr":196,"name":"Windspiel"},{"nr":197,"name":"Passatwind"},{"nr":198,"name":"Wellenbrecher"}],
    "Haus Reetzeit": [{"nr":95,"name":"Reetzeit 1"},{"nr":96,"name":"Reetzeit 2"}],
    "Haus Rosenberg Küstenharmonie": [{"nr":108,"name":"Kranichrast"},{"nr":158,"name":"Meeresbrise"},{"nr":123,"name":"Zeeskahn"}],
    "Haus Schwedengang": [{"nr":189,"name":"Mondzauber"},{"nr":188,"name":"Sonnenschein"}],
    "Haus Seeluft & Seestern": [{"nr":287,"name":"Seeluft"},{"nr":286,"name":"Seestern"}],
    "Haus Sterntaucher": [{"nr":231,"name":"Sterntaucher 1"},{"nr":232,"name":"Sterntaucher 2"},{"nr":233,"name":"Sterntaucher 3"},{"nr":234,"name":"Sterntaucher 4"},{"nr":235,"name":"Sterntaucher 5"},{"nr":236,"name":"Sterntaucher 6"}],
    "Haus Störtebeker": [{"nr":82,"name":"Küstenzauber 12a Whg.3"},{"nr":89,"name":"Schatzkiste 12/2"},{"nr":83,"name":"Störtebekerkoje 12/1"},{"nr":84,"name":"Störtebekerkoje 12/4"},{"nr":85,"name":"Störtebekerkoje 12/5"},{"nr":86,"name":"Störtebekerkoje 12/6"},{"nr":87,"name":"Störtebekerkoje 12a/4"},{"nr":88,"name":"Störtebekerkoje 12a/6"},{"nr":136,"name":"uns Leef HS 12a Whg.1"}],
    "Haus Tordalk": [{"nr":237,"name":"Tordalk 1"},{"nr":246,"name":"Tordalk 3"},{"nr":239,"name":"Tordalk 5"},{"nr":240,"name":"Tordalk 6"},{"nr":241,"name":"Tordalk 7"},{"nr":242,"name":"Tordalk 9"},{"nr":243,"name":"Tordalk 10"}],
    "Haus Windflüchter": [{"nr":179,"name":"Windflüchter 1"},{"nr":180,"name":"Windflüchter 2"},{"nr":181,"name":"Windflüchter 3"},{"nr":182,"name":"Windflüchter EG"}],
    "Haus Windwatt": [{"nr":177,"name":"Windwatt 2"},{"nr":178,"name":"Windwatt 4"}],
    "Haus Zur Heiderose": [{"nr":185,"name":"54 Grad"},{"nr":147,"name":"Künstlerkate"},{"nr":149,"name":"Mondmuschel"},{"nr":171,"name":"Strandglück"},{"nr":175,"name":"Wellenflüstern"},{"nr":122,"name":"Wildrose"}],
    "Residenz am Strand": [{"nr":151,"name":"Residenz 114"},{"nr":152,"name":"Residenz 120"},{"nr":153,"name":"Residenz 123"},{"nr":155,"name":"Residenz 232"},{"nr":156,"name":"Residenz 238"},{"nr":157,"name":"Residenz 242"},{"nr":159,"name":"Residenz 352"},{"nr":161,"name":"Residenz 567"},{"nr":162,"name":"Residenz 677"}],
    "Residenz Kormoran": [{"nr":276,"name":"Ankerzeit H7"},{"nr":238,"name":"Meerzeit D6"},{"nr":245,"name":"Windflüchter F5"}],
    "Speicherresidenz Barth": [{"nr":29,"name":"App. 1.5"},{"nr":30,"name":"App. 0.2"},{"nr":31,"name":"App. 0.3"},{"nr":32,"name":"App. 2.1"},{"nr":33,"name":"App. 2.2"},{"nr":34,"name":"App. 3.3"},{"nr":35,"name":"App. 3.4"},{"nr":36,"name":"App. 3.7"},{"nr":37,"name":"App. 3.11"},{"nr":38,"name":"App. 3.1"},{"nr":39,"name":"App. 4.10"},{"nr":40,"name":"App. 4.11"},{"nr":41,"name":"App. 4.7"},{"nr":42,"name":"App. 4.6"},{"nr":43,"name":"App. 5.6"},{"nr":44,"name":"App. 5.9"},{"nr":45,"name":"App. 5.10"},{"nr":46,"name":"App. 6.1"},{"nr":47,"name":"App. 4.4"},{"nr":48,"name":"App. 2.3"},{"nr":49,"name":"App. 5.11"},{"nr":50,"name":"App. 7.1"}],
    "Strandapartments Düne 7": [{"nr":70,"name":"Düne 7 Whg. 3"},{"nr":71,"name":"Düne 7 Whg. 5"},{"nr":72,"name":"Düne 7 Whg. 6"},{"nr":73,"name":"Düne 7 Whg. 7"},{"nr":74,"name":"Düne 7 Whg. 8"},{"nr":75,"name":"Düne 7 Whg. 9"},{"nr":76,"name":"Düne 7 Whg. 10"}],
    "Strandresort Fuhlendorf": [{"nr":204,"name":"Luv"},{"nr":284,"name":"Sonnenbirke"},{"nr":248,"name":"Sonnenzauber"}],
    "Villa Seeluft": [{"nr":100,"name":"Seeluft 3"},{"nr":150,"name":"Seeluft 8"}],
    "Villa Strandoase Rosenberg": [{"nr":211,"name":"Küstenkajüte Whg.1"},{"nr":212,"name":"Küstenkajüte Whg.2"},{"nr":213,"name":"Küstenkajüte Whg.5"},{"nr":214,"name":"Küstenkajüte Whg.7"}],
}
# Umgekehrtes Mapping: Objekt-Nr (str) → Hausname
_OBJEKT_ZU_HAUS = {
    str(obj["nr"]): haus
    for haus, objekte in APARTMENTHAUS_MAPPING.items()
    for obj in objekte
}

# ---------------------------------------------------------------------------
# Eigentümer-Mapping: Objekt-Nr (str) → {eigentuemer, provision_pct_excel}
# Wird aus Unterkuenfte_Provisionssatz.xlsx geladen (falls vorhanden)
# ---------------------------------------------------------------------------
EIGENTUEMER_EXCEL = os.path.join(SCRIPT_DIR, "Unterkuenfte_Provisionssatz.xlsx")

def _load_eigentuemer_mapping():
    """Lädt Eigentümer + Provisionssatz aus der Excel-Datei."""
    mapping = {}
    if not os.path.exists(EIGENTUEMER_EXCEL):
        return mapping
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EIGENTUEMER_EXCEL, read_only=True, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            eigentuemer = str(ws.cell(r, 1).value or "").strip()
            nr = str(ws.cell(r, 2).value or "").strip()
            prov_raw = str(ws.cell(r, 5).value or "").strip()
            if nr:
                mapping[nr] = {
                    "eigentuemer": eigentuemer,
                    "provision_pct_excel": prov_raw,
                }
        wb.close()
    except Exception:
        pass
    return mapping

_EIGENTUEMER_MAPPING = _load_eigentuemer_mapping()


def parse_german_number(s):
    """Parse a German-formatted number (e.g. '1.234,56') to float."""
    if not s or not s.strip():
        return 0.0
    s = s.strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def format_german_number(n, decimals=2):
    """Format a number in German style: 1.234,56"""
    if decimals == 0:
        formatted = f"{n:,.0f}"
    else:
        formatted = f"{n:,.{decimals}f}"
    # Swap . and , for German formatting
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted


def format_euro(n):
    """Format as Euro amount."""
    return f"{format_german_number(n)} \u20ac"


def read_objektstammdaten(xlsx_path):
    """Read property master data (Wohnfläche, Zimmer, Schlafplätze) from Objektstammdaten.xlsx."""
    if not HAS_OPENPYXL or not os.path.exists(xlsx_path):
        return None

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Objektstammdaten"]

    properties = []
    orte_summary = defaultdict(lambda: {"count": 0, "wohnflaeche": 0, "zimmer": 0,
                                         "schlafzimmer": 0, "badzimmer": 0, "max_personen": 0,
                                         "sauna": 0, "hund": 0, "kamin": 0})
    totals = {"count": 0, "wohnflaeche": 0, "zimmer": 0, "schlafzimmer": 0,
              "badzimmer": 0, "max_personen": 0, "sauna": 0, "hund": 0, "kamin": 0}

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[0]
        objnr = row[1]
        if not objnr or not name or not str(name).strip():
            continue  # Skip section headers

        ort = str(row[3]).strip() if row[3] else ""
        wf = row[4] if row[4] else 0
        zi = row[5] if row[5] else 0
        sz = row[6] if row[6] else 0
        bz = row[7] if row[7] else 0
        mp = row[8] if row[8] else 0
        sauna = 1 if row[11] and str(row[11]).strip() not in ("–", "-", "") else 0
        hund = 1 if row[12] and str(row[12]).strip() not in ("–", "-", "") else 0
        kamin = 1 if row[13] and str(row[13]).strip() not in ("–", "-", "") else 0

        properties.append({
            "name": str(name).strip(), "objnr": str(objnr).strip(),
            "ort": ort, "wohnflaeche": wf, "zimmer": zi,
            "schlafzimmer": sz, "badzimmer": bz, "max_personen": mp,
            "sauna": sauna, "hund": hund, "kamin": kamin,
        })

        totals["count"] += 1
        totals["wohnflaeche"] += wf
        totals["zimmer"] += zi
        totals["schlafzimmer"] += sz
        totals["badzimmer"] += bz
        totals["max_personen"] += mp
        totals["sauna"] += sauna
        totals["hund"] += hund
        totals["kamin"] += kamin

        o = orte_summary[ort]
        o["count"] += 1
        o["wohnflaeche"] += wf
        o["zimmer"] += zi
        o["schlafzimmer"] += sz
        o["badzimmer"] += bz
        o["max_personen"] += mp
        o["sauna"] += sauna
        o["hund"] += hund
        o["kamin"] += kamin

    wb.close()
    return {
        "properties": properties,
        "totals": totals,
        "orte": dict(orte_summary),
    }


def read_bookings(csv_path):
    """Read and parse bookings from the CSV file, including Zusatzkosten."""
    bookings = []
    zusatz_categories = []  # list of (col_index, category_name)

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        header1 = next(reader)

        # Zweite Zeile: entweder eine zweite Kopfzeile (Contao-UI-Export)
        # oder bereits die erste Datenzeile (fetch_contao_data.py-Export).
        # Wir legen sie in einen Puffer und verarbeiten sie ggf. als Datenzeile.
        try:
            header2 = next(reader)
        except StopIteration:
            header2 = None

        # Parse Zusatzkosten column pairs (Vermittler at odd idx, Eigentümer at idx+1)
        for i in range(17, len(header1), 2):
            name = header1[i].strip()
            if name:
                zusatz_categories.append((i, name))

        # Zeilen-Iterator: header2 voranstellen falls es eine Datenzeile ist
        def _rows():
            if header2 is not None:
                yield header2
            yield from reader

        for row in _rows():
            if len(row) < 17:
                continue
            status = row[9].strip()
            # "Buchung" = abgerechnet (Journal); weitere Status aus Salesbooking
            SKIP_STATUS = {"Storno", "Stornierung", "Stonierung", "cancelled", ""}
            if status in SKIP_STATUS:
                continue
            try:
                anreise = datetime.strptime(row[4].strip(), "%d.%m.%Y")
            except ValueError:
                continue

            # Parse Zusatzkosten for this row
            zusatz = {}
            zusatz_names = set()
            for col_idx, cat_name in zusatz_categories:
                v = parse_german_number(row[col_idx]) if col_idx < len(row) else 0.0
                e = parse_german_number(row[col_idx + 1]) if col_idx + 1 < len(row) else 0.0
                if v != 0 or e != 0:
                    zusatz[cat_name] = {"vermittler": v, "eigentuemer": e}
                    zusatz_names.add(cat_name)

            # Derive travel profiles from booked extras
            profiles = []
            if any("Kinderreisebett" in n for n in zusatz_names) or any("Kinderhochstuhl" in n for n in zusatz_names):
                profiles.append("Familie mit Kleinkind")
            if any("Hund" in n for n in zusatz_names):
                profiles.append("Urlaub mit Hund")
            if any("Mitreisende" in n and "Aufschlag" in n for n in zusatz_names):
                profiles.append("Größere Reisegruppe")
            if any("Sauna" in n or "Whirlpool" in n for n in zusatz_names):
                profiles.append("Wellness-Gäste")
            if any("Wallbox" in n for n in zusatz_names):
                profiles.append("E-Auto Reisende")
            if not profiles:
                profiles.append("Paare/Einzelreisende")

            nr_raw = row[0].strip()
            bookings.append({
                "objekt_nr": nr_raw.split()[0] if nr_raw else "",
                "unterkunft": row[1].strip(),
                "ort": row[2].strip(),
                "buchungsdatum": row[3].strip(),
                "anreise": anreise,
                "abreise_str": row[5].strip(),
                "naechte": parse_german_number(row[6]),
                "vorgang": row[7].strip(),
                "vertriebskanal": row[8].strip(),
                "reisepreis": parse_german_number(row[11]),
                "provision_pct": row[13].strip() if len(row) > 13 else "",
                "miete_gesamt": parse_german_number(row[14]),
                "miete_vermittler": parse_german_number(row[15]),
                "miete_eigentuemer": parse_german_number(row[16]),
                "zusatzkosten": zusatz,
                "profiles": profiles,
            })
    return bookings


def compute_data(bookings):
    """Compute all aggregated data for the dashboard."""
    import datetime as _dt
    _now = _dt.datetime.now()
    _current_year = _now.year
    _next_year = _now.year + 1

    # --- Per-year KPIs ---
    # Always include last 2 years + current year + next year, even if no bookings yet
    _booking_years = set(b["anreise"].year for b in bookings)
    _booking_years.add(_now.year - 2)
    _booking_years.add(_now.year - 1)
    _booking_years.add(_current_year)
    _booking_years.add(_next_year)
    years = sorted(_booking_years)
    kpis = {}
    for y in years:
        yb = [b for b in bookings if b["anreise"].year == y]
        kpis[y] = {
            "buchungen": len(yb),
            "naechte": sum(b["naechte"] for b in yb),
            "reisepreis": sum(b["reisepreis"] for b in yb),
            "miete_gesamt": sum(b["miete_gesamt"] for b in yb),
            "miete_vermittler": sum(b["miete_vermittler"] for b in yb),
            "miete_eigentuemer": sum(b["miete_eigentuemer"] for b in yb),
        }

    # --- Monthly nights per year (for line chart) ---
    monthly = defaultdict(lambda: defaultdict(float))
    for b in bookings:
        y = b["anreise"].year
        m = b["anreise"].month
        monthly[y][m] += b["naechte"]

    monthly_data = {}
    for y in years:
        monthly_data[y] = [monthly[y].get(m, 0) for m in range(1, 13)]

    # --- Monthly bookings count per year (for comparison table) ---
    monthly_count = defaultdict(lambda: defaultdict(int))
    for b in bookings:
        monthly_count[b["anreise"].year][b["anreise"].month] += 1

    monthly_count_data = {}
    for y in years:
        monthly_count_data[y] = [monthly_count[y].get(m, 0) for m in range(1, 13)]

    # --- Top 10 properties by booking count ---
    prop_counts = defaultdict(int)
    for b in bookings:
        prop_counts[b["unterkunft"]] += 1
    top_props = sorted(prop_counts.items(), key=lambda x: -x[1])[:10]

    # --- Reiseprofile / Zielgruppen ---
    profile_order = ["Paare/Einzelreisende", "Urlaub mit Hund", "Größere Reisegruppe",
                     "Familie mit Kleinkind", "Wellness-Gäste", "E-Auto Reisende"]
    profile_colors = {
        "Paare/Einzelreisende": "#0066cc",
        "Urlaub mit Hund": "#ff6b6b",
        "Größere Reisegruppe": "#ffa500",
        "Familie mit Kleinkind": "#4ecdc4",
        "Wellness-Gäste": "#aa96da",
        "E-Auto Reisende": "#95e1d3",
    }

    # Per year totals
    profiles_by_year = defaultdict(lambda: defaultdict(int))
    for b in bookings:
        y = b["anreise"].year
        for p in b["profiles"]:
            profiles_by_year[y][p] += 1

    # Per property: profile breakdown (all years combined + per year)
    prop_profiles = defaultdict(lambda: {"total": defaultdict(int), "per_year": defaultdict(lambda: defaultdict(int)), "buchungen": 0})
    for b in bookings:
        y = b["anreise"].year
        name = b["unterkunft"]
        prop_profiles[name]["buchungen"] += 1
        for p in b["profiles"]:
            prop_profiles[name]["total"][p] += 1
            prop_profiles[name]["per_year"][y][p] += 1

    # --- Sales channels ---
    channel_counts = defaultdict(int)
    for b in bookings:
        channel_counts[b["vertriebskanal"]] += 1
    channels_sorted = sorted(channel_counts.items(), key=lambda x: -x[1])

    # --- Sales channels per year (Ostseeliebe vs. Portalbuchungen grouping) ---
    OSTSEELIEBE_CHANNELS = {
        "Webseite", "Telefon", "Email", "E-Mail", "Büro", "Buchungsanfrage",
        "Online-Reservierung", "Whatsapp", "Eigentümer-Login", "Kundenangebot",
        "Umbuchungen", "Buchung durch Eigentümer", "Newsletter",
        "vorherige Agentur", "Wiederbucher",
    }
    channels_by_year = {}
    for y in years:
        yb = [b for b in bookings if b["anreise"].year == y]
        year_total = len(yb)
        raw = defaultdict(int)
        for b in yb:
            raw[b["vertriebskanal"]] += 1

        # Portalprovision pro Kanal: Reisepreis × Provision% (z.B. "17%" → 0.17)
        prov_by_channel = defaultdict(float)
        for b in yb:
            raw_pct = b["provision_pct"].replace("%", "").replace(",", ".").strip()
            try:
                pct = float(raw_pct) / 100.0
            except (ValueError, AttributeError):
                pct = 0.0
            prov_by_channel[b["vertriebskanal"]] += b["reisepreis"] * pct

        # Classify each channel
        ostl_sub = []
        portal_sub = []
        for k, v in sorted(raw.items(), key=lambda x: -x[1]):
            if not k:
                continue
            entry = {"name": k, "count": v, "pct": round(100 * v / year_total, 1) if year_total else 0,
                     "provision": round(prov_by_channel.get(k, 0), 2)}
            if k in OSTSEELIEBE_CHANNELS:
                ostl_sub.append(entry)
            else:
                portal_sub.append(entry)

        grouped = []
        ostl_total = sum(e["count"] for e in ostl_sub)
        if ostl_total > 0:
            grouped.append({"name": "Ostseeliebe", "count": ostl_total,
                            "pct": round(100 * ostl_total / year_total, 1) if year_total else 0,
                            "provision": 0,  # keine Portalprovision für eigene Kanäle
                            "sub": ostl_sub})
        portal_total = sum(e["count"] for e in portal_sub)
        portal_prov_total = round(sum(e["provision"] for e in portal_sub), 2)
        if portal_total > 0:
            grouped.append({"name": "Portalbuchungen", "count": portal_total,
                            "pct": round(100 * portal_total / year_total, 1) if year_total else 0,
                            "provision": portal_prov_total,
                            "sub": portal_sub})
        channels_by_year[y] = {"total": year_total, "channels": grouped,
                                "portal_provision": portal_prov_total}

    # --- Locations – Umsatz + Buchungen pro Jahr ---
    ort_data = defaultdict(lambda: {"buchungen": defaultdict(int), "umsatz": defaultdict(float)})
    for b in bookings:
        ort = b["ort"]
        y = b["anreise"].year
        ort_data[ort]["buchungen"][y] += 1
        ort_data[ort]["umsatz"][y] += b["reisepreis"]

    # orte_sorted: Liste mit vollständigen Jahresdaten
    orte_sorted = []
    for ort, od in ort_data.items():
        total_umsatz = sum(od["umsatz"].values())
        total_buch   = sum(od["buchungen"].values())
        orte_sorted.append({
            "ort": ort,
            "total_umsatz": total_umsatz,
            "total_buchungen": total_buch,
            "umsatz_per_year": dict(od["umsatz"]),
            "buchungen_per_year": dict(od["buchungen"]),
        })
    orte_sorted.sort(key=lambda x: -x["total_umsatz"])

    # --- Zusatzkosten aggregation ---
    # Structure: {category: {year: {vermittler, eigentuemer, count}}}
    zusatz_by_cat_year = defaultdict(lambda: defaultdict(lambda: {"vermittler": 0.0, "eigentuemer": 0.0, "count": 0}))
    for b in bookings:
        y = b["anreise"].year
        for cat_name, vals in b["zusatzkosten"].items():
            zusatz_by_cat_year[cat_name][y]["vermittler"] += vals["vermittler"]
            zusatz_by_cat_year[cat_name][y]["eigentuemer"] += vals["eigentuemer"]
            zusatz_by_cat_year[cat_name][y]["count"] += 1

    # Build sorted list by absolute total across all years
    zusatz_sorted = []
    for cat, year_data in zusatz_by_cat_year.items():
        total_v = sum(d["vermittler"] for d in year_data.values())
        total_e = sum(d["eigentuemer"] for d in year_data.values())
        total_c = sum(d["count"] for d in year_data.values())
        zusatz_sorted.append({
            "name": cat,
            "total": total_v + total_e,
            "vermittler": total_v,
            "eigentuemer": total_e,
            "count": total_c,
            "per_year": {y: {"vermittler": year_data[y]["vermittler"],
                             "eigentuemer": year_data[y]["eigentuemer"],
                             "gesamt": year_data[y]["vermittler"] + year_data[y]["eigentuemer"],
                             "count": year_data[y]["count"]} for y in years}
        })
    zusatz_sorted.sort(key=lambda x: abs(x["total"]), reverse=True)

    # Per-year Zusatzkosten totals (for Übersicht summary)
    zusatz_year_totals = {}
    for y in years:
        v_sum = sum(z["per_year"][y]["vermittler"] for z in zusatz_sorted)
        e_sum = sum(z["per_year"][y]["eigentuemer"] for z in zusatz_sorted)
        zusatz_year_totals[y] = {"vermittler": v_sum, "eigentuemer": e_sum, "gesamt": v_sum + e_sum}

    # --- Per-property detailed data ---
    property_data = {}  # {name: {...}}
    prop_bookings = defaultdict(list)
    for b in bookings:
        prop_bookings[b["unterkunft"]].append(b)

    for prop_name, pbs in sorted(prop_bookings.items()):
        ort = pbs[0]["ort"] if pbs else ""
        obj_nr = pbs[0]["objekt_nr"] if pbs else ""
        em = _EIGENTUEMER_MAPPING.get(obj_nr, {})
        eigentuemer = em.get("eigentuemer", "")
        provision_pct_excel = em.get("provision_pct_excel", "")
        p_years = {}
        for y in years:
            ybs = [b for b in pbs if b["anreise"].year == y]
            if not ybs:
                p_years[y] = {
                    "buchungen": 0, "naechte": 0, "reisepreis": 0, "miete_gesamt": 0,
                    "miete_vermittler": 0, "miete_eigentuemer": 0, "provision_pct": 0,
                    "avg_preis_nacht": 0, "avg_aufenthalt": 0,
                    "belegung_pct": 0, "channels": {}, "zusatzkosten": {}
                }
                continue
            n_buch = len(ybs)
            n_naechte = sum(b["naechte"] for b in ybs)
            rp = sum(b["reisepreis"] for b in ybs)
            mg = sum(b["miete_gesamt"] for b in ybs)
            mv = sum(b["miete_vermittler"] for b in ybs)
            me = sum(b["miete_eigentuemer"] for b in ybs)
            prov_pct = round(mv / mg * 100, 1) if mg > 0 else 0
            avg_pn = rp / n_naechte if n_naechte > 0 else 0
            avg_auf = n_naechte / n_buch if n_buch > 0 else 0
            beleg = round(n_naechte / 365 * 100, 1) if n_naechte > 0 else 0

            # Channels for this property/year
            ch = defaultdict(int)
            for b in ybs:
                ch[b["vertriebskanal"]] += 1
            ch_sorted = sorted(ch.items(), key=lambda x: -x[1])

            # Zusatzkosten for this property/year
            zk = defaultdict(lambda: {"vermittler": 0.0, "eigentuemer": 0.0, "count": 0})
            for b in ybs:
                for cat, vals in b["zusatzkosten"].items():
                    zk[cat]["vermittler"] += vals["vermittler"]
                    zk[cat]["eigentuemer"] += vals["eigentuemer"]
                    zk[cat]["count"] += 1
            zk_list = [{"name": k, "vermittler": v["vermittler"], "eigentuemer": v["eigentuemer"],
                         "gesamt": v["vermittler"] + v["eigentuemer"], "count": v["count"]}
                        for k, v in zk.items() if v["vermittler"] != 0 or v["eigentuemer"] != 0]
            zk_list.sort(key=lambda x: abs(x["gesamt"]), reverse=True)

            p_years[y] = {
                "buchungen": n_buch, "naechte": n_naechte, "reisepreis": rp,
                "miete_gesamt": mg, "miete_vermittler": mv, "miete_eigentuemer": me,
                "provision_pct": prov_pct,
                "avg_preis_nacht": round(avg_pn, 2), "avg_aufenthalt": round(avg_auf, 1),
                "belegung_pct": beleg,
                "channels": ch_sorted, "zusatzkosten": zk_list
            }

        property_data[prop_name] = {
            "ort": ort, "years": p_years,
            "objekt_nr": obj_nr,
            "eigentuemer": eigentuemer,
            "provision_pct_excel": provision_pct_excel,
        }

    # --- Provision summary per property (for Provisionen tab) ---
    provision_by_prop = {}
    for prop_name, pd in sorted(property_data.items()):
        prop_prov = {}
        # Vertragssatz aus Excel als Dezimalzahl (z.B. "17%" → 0.17)
        _pct_excel_str = pd.get("provision_pct_excel", "")
        try:
            _contracted_rate = float(_pct_excel_str.replace("%", "").replace(",", ".").strip()) / 100
        except (ValueError, AttributeError):
            _contracted_rate = 0.0
        for y in years:
            yd = pd["years"].get(y, {})
            if yd.get("buchungen", 0) == 0:
                continue
            zk_verm = sum(z["vermittler"] for z in yd.get("zusatzkosten", []))
            mg = yd["miete_gesamt"]
            # Provision laut Vertrag = Miete gesamt × Vertragssatz
            prov_laut_vertrag = round(mg * _contracted_rate, 2) if _contracted_rate else None
            prop_prov[y] = {
                "buchungen": yd["buchungen"],
                "miete_gesamt": mg,
                "miete_vermittler": yd["miete_vermittler"],
                "zusatz_vermittler": round(zk_verm, 2),
                "provision_gesamt": round(yd["miete_vermittler"], 2),  # Contao-Wert
                "provision_laut_vertrag": prov_laut_vertrag,           # Option A: MG × Vertragssatz
                "provision_pct": yd["provision_pct"],
            }
        if prop_prov:
            provision_by_prop[prop_name] = {
                "ort": pd["ort"],
                "objekt_nr": pd.get("objekt_nr", ""),
                "eigentuemer": pd.get("eigentuemer", ""),
                "provision_pct_excel": pd.get("provision_pct_excel", ""),
                "years": prop_prov,
            }

    # --- Apartmenthaus aggregation ---
    # {haus_name: {year: {unit_name: {buchungen, umsatz}}}}
    haus_year_data = {}
    for haus_name, objekte in APARTMENTHAUS_MAPPING.items():
        if not objekte:
            continue
        haus_year_data[haus_name] = {
            y: {obj["name"]: {"buchungen": 0, "umsatz": 0.0, "miete_eigentuemer": 0.0} for obj in objekte}
            for y in years
        }
    for b in bookings:
        nr_str = b["objekt_nr"]
        haus_name = _OBJEKT_ZU_HAUS.get(nr_str)
        if not haus_name or haus_name not in haus_year_data:
            continue
        y = b["anreise"].year
        if y not in years:
            continue
        for obj in APARTMENTHAUS_MAPPING[haus_name]:
            if str(obj["nr"]) == nr_str:
                unit = obj["name"]
                haus_year_data[haus_name][y][unit]["buchungen"] += 1
                haus_year_data[haus_name][y][unit]["umsatz"] += b["reisepreis"]
                haus_year_data[haus_name][y][unit]["miete_eigentuemer"] += b["miete_eigentuemer"]
                break

    # --- Provision aggregation per Eigentümer ---
    # {eigentuemer: {year: {buchungen, miete_gesamt, miete_vermittler, provision_gesamt}}}
    provision_by_eigentuemer = {}
    # Sammle zuerst alle Vertragssätze je Eigentümer (über alle Jahre hinweg)
    _eg_rates = {}  # eg → set of provision_pct_excel strings
    for prop_name, pprov in provision_by_prop.items():
        eg = pprov.get("eigentuemer", "") or "Unbekannt"
        rate = pprov.get("provision_pct_excel", "")
        if eg not in _eg_rates:
            _eg_rates[eg] = set()
        if rate:
            _eg_rates[eg].add(rate)

    for prop_name, pprov in provision_by_prop.items():
        eg = pprov.get("eigentuemer", "") or "Unbekannt"
        if eg not in provision_by_eigentuemer:
            provision_by_eigentuemer[eg] = {}
        for y, py in pprov["years"].items():
            if y not in provision_by_eigentuemer[eg]:
                provision_by_eigentuemer[eg][y] = {
                    "buchungen": 0, "miete_gesamt": 0.0,
                    "miete_vermittler": 0.0, "provision_gesamt": 0.0,
                    "provision_laut_vertrag": 0.0,
                    "unterkuenfte": [],
                    "provision_saetze": sorted(_eg_rates.get(eg, set())),
                }
            provision_by_eigentuemer[eg][y]["buchungen"] += py["buchungen"]
            provision_by_eigentuemer[eg][y]["miete_gesamt"] += py["miete_gesamt"]
            provision_by_eigentuemer[eg][y]["miete_vermittler"] += py["miete_vermittler"]
            provision_by_eigentuemer[eg][y]["provision_gesamt"] += py["provision_gesamt"]
            if py.get("provision_laut_vertrag") is not None:
                provision_by_eigentuemer[eg][y]["provision_laut_vertrag"] += py["provision_laut_vertrag"]
            provision_by_eigentuemer[eg][y]["unterkuenfte"].append(prop_name)

    return {
        "years": years,
        "kpis": kpis,
        "monthly_data": monthly_data,
        "monthly_count_data": monthly_count_data,
        "top_props": top_props,
        "channels": channels_sorted,
        "channels_by_year": channels_by_year,
        "orte": orte_sorted,
        "zusatz_sorted": zusatz_sorted,
        "zusatz_year_totals": zusatz_year_totals,
        "property_data": property_data,
        "provision_by_prop": provision_by_prop,
        "provision_by_eigentuemer": provision_by_eigentuemer,
        "profile_order": profile_order,
        "profile_colors": profile_colors,
        "profiles_by_year": profiles_by_year,
        "prop_profiles": prop_profiles,
        "haus_year_data": haus_year_data,
    }


def _build_orte_tab(orte, years, current_year):
    """Baut den Orte-Tab mit Umsatz pro Jahr, %-Anteil und Wachstum."""
    # Nur Hauptjahre (ab 2020), aktuelles Jahr vorne
    MAIN_FROM = 2020
    display_years = sorted(
        [y for y in years if y >= MAIN_FROM],
        key=lambda y: (0 if y == current_year else 1, -y)
    )

    # Jahrestotale für %-Berechnung
    year_totals = {}
    for y in display_years:
        year_totals[y] = sum(o["umsatz_per_year"].get(y, 0) for o in orte)

    # Tabellenkopf
    year_headers = ""
    for y in display_years:
        is_cur = y == current_year
        bg = "#0066cc" if is_cur else "#4a90d9"
        year_headers += (
            f'<th class="num" colspan="2" style="background:{bg};color:#fff;'
            f'font-weight:700;text-align:center;">'
            f'{"▶ " if is_cur else ""}{y}</th>'
            f'<th class="num" style="background:{"#1a7de0" if is_cur else "#5fa3e8"};'
            f'color:#fff;font-size:11px;font-weight:500;">Wachstum</th>'
        )

    sub_headers = ""
    for y in display_years:
        is_cur = y == current_year
        bg2 = "#1a7de0" if is_cur else "#5fa3e8"
        sub_headers += (
            f'<th class="num" style="background:{bg2};color:#fff;font-size:11px;font-weight:500;">Umsatz</th>'
            f'<th class="num" style="background:{bg2};color:#fff;font-size:11px;font-weight:500;">Anteil</th>'
            f'<th class="num" style="background:{bg2};color:#fff;font-size:11px;font-weight:500;">vs. Vorjahr</th>'
        )

    # Zeilen – nach aktuellem Jahr sortiert
    orte_sorted_cur = sorted(orte, key=lambda o: -o["umsatz_per_year"].get(current_year, 0))

    rows_html = ""
    for o in orte_sorted_cur:
        cells = ""
        for idx, y in enumerate(display_years):
            umsatz = o["umsatz_per_year"].get(y, 0)
            total_y = year_totals.get(y, 1) or 1
            anteil = umsatz / total_y * 100 if umsatz else 0

            # Wachstum vs. Vorjahr
            prev_y = display_years[idx + 1] if idx + 1 < len(display_years) else None
            if prev_y:
                prev = o["umsatz_per_year"].get(prev_y, 0)
                if prev > 0:
                    wachstum = (umsatz - prev) / prev * 100
                    if wachstum > 0:
                        wstr = f'<span style="color:#28a745;font-weight:600">▲ {wachstum:+.1f}%</span>'
                    elif wachstum < 0:
                        wstr = f'<span style="color:#dc3545;font-weight:600">▼ {wachstum:.1f}%</span>'
                    else:
                        wstr = '<span style="color:#888">± 0%</span>'
                elif umsatz > 0:
                    wstr = '<span style="color:#28a745;font-size:11px">neu</span>'
                else:
                    wstr = '<span style="color:#ccc">–</span>'
            else:
                wstr = '<span style="color:#ccc">–</span>'

            is_cur = y == current_year
            bg_cell = "background:#f0f7ff;" if is_cur else ""
            fw = "font-weight:600;" if is_cur else ""
            umsatz_str = format_euro(umsatz) if umsatz else '<span style="color:#ccc">–</span>'
            anteil_str = f'{anteil:.1f} %' if umsatz else '<span style="color:#ccc">–</span>'
            cells += (
                f'<td class="num" style="{bg_cell}{fw}">{umsatz_str}</td>'
                f'<td class="num" style="{bg_cell}color:#666;font-size:12px;">{anteil_str}</td>'
                f'<td class="num" style="{bg_cell}font-size:12px;">{wstr}</td>'
            )

        rows_html += f'<tr><td style="font-weight:500">{o["ort"]}</td>{cells}</tr>\n'

    # Summenzeile
    sum_cells = ""
    for idx, y in enumerate(display_years):
        total_y = year_totals.get(y, 0)
        prev_y = display_years[idx + 1] if idx + 1 < len(display_years) else None
        if prev_y and year_totals.get(prev_y, 0) > 0:
            wachstum = (total_y - year_totals[prev_y]) / year_totals[prev_y] * 100
            if wachstum > 0:
                wstr = f'<span style="color:#28a745;font-weight:600">▲ {wachstum:+.1f}%</span>'
            else:
                wstr = f'<span style="color:#dc3545;font-weight:600">▼ {wachstum:.1f}%</span>'
        else:
            wstr = '<span style="color:#ccc">–</span>'
        is_cur = y == current_year
        bg_cell = "background:#e8f4ff;" if is_cur else "background:#f5f5f5;"
        sum_cells += (
            f'<td class="num" style="{bg_cell}font-weight:700">{format_euro(total_y)}</td>'
            f'<td class="num" style="{bg_cell}font-weight:700">100 %</td>'
            f'<td class="num" style="{bg_cell}font-size:12px;">{wstr}</td>'
        )

    return f'''<div class="chart-container">
        <h3>&#127968; Umsatz nach Ort</h3>
        <p style="color:#666;font-size:13px;margin-bottom:16px;">
            Reisepreis je Ort und Jahr &mdash; sortiert nach Umsatz {current_year} &mdash;
            Wachstum jeweils vs. Vorjahr.
        </p>
        <div style="overflow-x:auto;">
        <table class="prov-table">
            <thead>
                <tr>
                    <th rowspan="2" style="vertical-align:bottom;">Ort</th>
                    {year_headers}
                </tr>
                <tr>{sub_headers}</tr>
            </thead>
            <tbody>
                {rows_html}
                <tr class="prov-total">
                    <td><strong>GESAMT</strong></td>
                    {sum_cells}
                </tr>
            </tbody>
        </table>
        </div>
    </div>'''


def generate_html(data):
    """Generate the complete dashboard HTML."""
    import datetime as _dt
    _current_year = _dt.datetime.now().year
    _next_year = _current_year + 1
    _all_years = data["years"]
    # Hauptansicht: ab 2024, aktuelles Jahr zuerst, dann absteigend
    # Archiv: alles vor 2024, absteigend (neuestes zuerst)
    MAIN_FROM = 2024
    years = sorted(
        [y for y in _all_years if y >= MAIN_FROM],
        key=lambda y: (0 if y == _current_year else 1, -y)
    )
    years_archive = sorted([y for y in _all_years if y < MAIN_FROM], reverse=True)
    kpis = data["kpis"]
    monthly_data = data["monthly_data"]
    monthly_count_data = data["monthly_count_data"]
    top_props = data["top_props"]
    channels = data["channels"]
    channels_by_year = data["channels_by_year"]
    orte = data["orte"]
    zusatz_sorted = data["zusatz_sorted"]
    zusatz_year_totals = data["zusatz_year_totals"]
    property_data = data["property_data"]
    provision_by_prop = data["provision_by_prop"]
    provision_by_eigentuemer = data.get("provision_by_eigentuemer", {})
    profile_order = data["profile_order"]
    profile_colors = data["profile_colors"]
    profiles_by_year = data["profiles_by_year"]
    prop_profiles = data["prop_profiles"]
    haus_year_data = data.get("haus_year_data", {})
    stammdaten = data.get("stammdaten")

    update_date = datetime.now().strftime("%d.%m.%Y %H:%M")
    month_names = [
        "Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
        "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"
    ]
    colors = [
        "#0066cc", "#00aaaa", "#ff6b6b", "#ffa500", "#4ecdc4",
        "#95e1d3", "#f38181", "#aa96da", "#fcbad3", "#a8dadc"
    ]

    # --- Build Bestandskennzahlen HTML (from Objektstammdaten) ---
    bestand_html = ""
    if stammdaten:
        t = stammdaten["totals"]
        orte_stamm = stammdaten["orte"]

        # Ort rows for the table
        ort_rows = ""
        for ort_name in sorted(orte_stamm.keys()):
            o = orte_stamm[ort_name]
            ort_rows += f'''<tr>
                <td>{ort_name}</td>
                <td class="zk-num">{o["count"]}</td>
                <td class="zk-num">{format_german_number(o["wohnflaeche"], 0)} m²</td>
                <td class="zk-num">{format_german_number(o["zimmer"], 0)}</td>
                <td class="zk-num">{o["schlafzimmer"]}</td>
                <td class="zk-num">{o["badzimmer"]}</td>
                <td class="zk-num">{o["max_personen"]}</td>
                <td class="zk-num">{o["sauna"]}</td>
                <td class="zk-num">{o["hund"]}</td>
                <td class="zk-num">{o["kamin"]}</td>
            </tr>'''

        bestand_html = f'''
        <div class="year-section">
            <h3 class="year-title">Bestand – Buchbare Unterkünfte (Online)</h3>
            <div class="kpi-grid">
                <div class="kpi-card" style="border-left:4px solid #28a745">
                    <div class="kpi-label">Unterkünfte</div>
                    <div class="kpi-value" style="color:#28a745">{t["count"]}</div>
                </div>
                <div class="kpi-card" style="border-left:4px solid #28a745">
                    <div class="kpi-label">Wohnfläche gesamt</div>
                    <div class="kpi-value" style="color:#28a745">{format_german_number(t["wohnflaeche"], 0)} m²</div>
                </div>
                <div class="kpi-card" style="border-left:4px solid #28a745">
                    <div class="kpi-label">Schlafzimmer</div>
                    <div class="kpi-value" style="color:#28a745">{t["schlafzimmer"]}</div>
                </div>
                <div class="kpi-card" style="border-left:4px solid #28a745">
                    <div class="kpi-label">Schlafplätze (max.)</div>
                    <div class="kpi-value" style="color:#28a745">{t["max_personen"]}</div>
                </div>
                <div class="kpi-card" style="border-left:4px solid #28a745">
                    <div class="kpi-label">Zimmer gesamt</div>
                    <div class="kpi-value" style="color:#28a745">{format_german_number(t["zimmer"], 0)}</div>
                </div>
            </div>
            <div class="kpi-grid" style="grid-template-columns:repeat(3,1fr);margin-top:8px">
                <div class="kpi-card" style="border-left:4px solid #aa96da">
                    <div class="kpi-label">mit Sauna</div>
                    <div class="kpi-value" style="color:#aa96da">{t["sauna"]}</div>
                </div>
                <div class="kpi-card" style="border-left:4px solid #ff6b6b">
                    <div class="kpi-label">Hund erlaubt</div>
                    <div class="kpi-value" style="color:#ff6b6b">{t["hund"]}</div>
                </div>
                <div class="kpi-card" style="border-left:4px solid #ffa500">
                    <div class="kpi-label">mit Kamin</div>
                    <div class="kpi-value" style="color:#ffa500">{t["kamin"]}</div>
                </div>
            </div>
            <div class="zusatz-summary" style="margin-top:12px">
                <table class="zusatz-table">
                    <thead>
                        <tr>
                            <th>Ort</th>
                            <th class="zk-num">Objekte</th>
                            <th class="zk-num">Wohnfl.</th>
                            <th class="zk-num">Zimmer</th>
                            <th class="zk-num">Schlafz.</th>
                            <th class="zk-num">Bäder</th>
                            <th class="zk-num">Schlafpl.</th>
                            <th class="zk-num">Sauna</th>
                            <th class="zk-num">Hund</th>
                            <th class="zk-num">Kamin</th>
                        </tr>
                    </thead>
                    <tbody>
                        {ort_rows}
                        <tr class="zk-total">
                            <td><strong>Gesamt</strong></td>
                            <td class="zk-num"><strong>{t["count"]}</strong></td>
                            <td class="zk-num"><strong>{format_german_number(t["wohnflaeche"], 0)} m²</strong></td>
                            <td class="zk-num"><strong>{format_german_number(t["zimmer"], 0)}</strong></td>
                            <td class="zk-num"><strong>{t["schlafzimmer"]}</strong></td>
                            <td class="zk-num"><strong>{t["badzimmer"]}</strong></td>
                            <td class="zk-num"><strong>{t["max_personen"]}</strong></td>
                            <td class="zk-num"><strong>{t["sauna"]}</strong></td>
                            <td class="zk-num"><strong>{t["hund"]}</strong></td>
                            <td class="zk-num"><strong>{t["kamin"]}</strong></td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>'''

    # --- Build KPI cards HTML (with Zusatzkosten summary) ---
    # Determine top Zusatzkosten categories (those with abs total > 100)
    top_zusatz = [z for z in zusatz_sorted if abs(z["total"]) > 100]

    kpi_html_parts = []
    # years ist bereits: aktuelles Jahr zuerst, dann absteigend
    current_year = _current_year
    years_sorted = years  # Reihenfolge wird global in years festgelegt
    for y in years_sorted:
        k = kpis[y]
        zt = zusatz_year_totals[y]

        # Build Zusatzkosten mini-table for this year
        zusatz_rows = ""
        for z in top_zusatz:
            zy = z["per_year"][y]
            if zy["gesamt"] == 0 and zy["vermittler"] == 0 and zy["eigentuemer"] == 0:
                continue
            zusatz_rows += f'''<tr>
                <td>{z["name"]}</td>
                <td class="zk-num">{format_euro(zy["gesamt"])}</td>
                <td class="zk-num zk-sub">{format_euro(zy["vermittler"])}</td>
                <td class="zk-num zk-sub">{format_euro(zy["eigentuemer"])}</td>
            </tr>'''

        zusatz_total_row = f'''<tr class="zk-total">
            <td><strong>Summe Zusatzkosten</strong></td>
            <td class="zk-num"><strong>{format_euro(zt["gesamt"])}</strong></td>
            <td class="zk-num zk-sub"><strong>{format_euro(zt["vermittler"])}</strong></td>
            <td class="zk-num zk-sub"><strong>{format_euro(zt["eigentuemer"])}</strong></td>
        </tr>'''

        kpi_html_parts.append(f'''
        <div class="year-section">
            <h3 class="year-title">{y}</h3>
            <div class="kpi-grid">
                <div class="kpi-card">
                    <div class="kpi-label">Buchungen</div>
                    <div class="kpi-value">{k["buchungen"]}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-label">N\u00e4chte</div>
                    <div class="kpi-value">{format_german_number(k["naechte"], 0)}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-label">Reisepreis</div>
                    <div class="kpi-value">{format_euro(k["reisepreis"])}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-label">Miete gesamt</div>
                    <div class="kpi-value">{format_euro(k["miete_gesamt"])}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-label">Miete Eigent\u00fcmer</div>
                    <div class="kpi-value">{format_euro(k["miete_eigentuemer"])}</div>
                </div>
            </div>
            <div class="zusatz-summary">
                <table class="zusatz-table">
                    <thead>
                        <tr>
                            <th>Zusatzkosten</th>
                            <th class="zk-num">Gesamt</th>
                            <th class="zk-num zk-sub">Vermittler</th>
                            <th class="zk-num zk-sub">Eigent\u00fcmer</th>
                        </tr>
                    </thead>
                    <tbody>
                        {zusatz_rows}
                        {zusatz_total_row}
                    </tbody>
                </table>
            </div>
        </div>''')
    kpi_html = "\n".join(kpi_html_parts)

    # --- Archiv-Zusammenfassung für Übersicht (2017–2023) ---
    archive_kpi_rows = ""
    for y in years_archive:
        k = kpis.get(y, {})
        if not k or k.get("buchungen", 0) == 0:
            continue
        archive_kpi_rows += f'''
        <tr>
            <td><strong>{y}</strong></td>
            <td class="num">{format_german_number(k.get("buchungen", 0), 0)}</td>
            <td class="num">{format_german_number(k.get("naechte", 0), 0)}</td>
            <td class="num">{format_euro(k.get("reisepreis", 0))}</td>
            <td class="num">{format_euro(k.get("miete_gesamt", 0))}</td>
            <td class="num">{format_euro(k.get("miete_eigentuemer", 0))}</td>
        </tr>'''

    if archive_kpi_rows:
        archive_min = min(y for y in years_archive if kpis.get(y, {}).get("buchungen", 0) > 0)
        archive_max = max(y for y in years_archive if kpis.get(y, {}).get("buchungen", 0) > 0)
        kpi_html += f'''
    <div style="margin-top:24px;">
        <div onclick="var c=document.getElementById('archiv-kpi');var a=document.getElementById('archiv-kpi-arrow');c.style.display=c.style.display==='none'?'block':'none';a.textContent=c.style.display==='none'?'▶':'▼';"
             style="cursor:pointer;background:#f5f5f5;border:1px solid #ddd;border-radius:8px;padding:12px 18px;display:flex;align-items:center;gap:12px;user-select:none;">
            <span id="archiv-kpi-arrow" style="font-size:16px;color:#888;">▶</span>
            <span style="font-weight:600;color:#666;">Archiv {archive_min}–{archive_max}</span>
        </div>
        <div id="archiv-kpi" style="display:none;margin-top:8px;overflow-x:auto;">
            <table class="prov-table">
                <thead>
                    <tr>
                        <th>Jahr</th>
                        <th class="num">Buchungen</th>
                        <th class="num">Nächte</th>
                        <th class="num">Reisepreis</th>
                        <th class="num">Miete gesamt</th>
                        <th class="num">Miete Eigentümer</th>
                    </tr>
                </thead>
                <tbody>{archive_kpi_rows}</tbody>
            </table>
        </div>
    </div>'''

    # --- Build comparison table HTML ---
    table_header = "<tr><th>Monat</th>" + "".join(f"<th>{y}</th>" for y in years) + "</tr>"
    table_rows = []
    for m_idx, m_name in enumerate(month_names):
        row = f"<tr><td>{m_name}</td>"
        for y in years:
            val = monthly_count_data[y][m_idx]
            row += f"<td>{val}</td>"
        row += "</tr>"
        table_rows.append(row)
    # Totals row
    totals_row = "<tr class='total-row'><td><strong>Gesamt</strong></td>"
    for y in years:
        totals_row += f"<td><strong>{sum(monthly_count_data[y])}</strong></td>"
    totals_row += "</tr>"
    table_rows.append(totals_row)
    comparison_table = f"<table class='comparison-table'><thead>{table_header}</thead><tbody>{''.join(table_rows)}</tbody></table>"

    # --- Build Zusatzkosten detail tab HTML ---
    # Full detail table with all categories and years
    zusatz_detail_header = "<tr><th>Kategorie</th><th class='zk-num'>Anzahl</th>"
    for y in years_sorted:
        zusatz_detail_header += f"<th class='zk-num'>{y} Ges.</th><th class='zk-num zk-sub'>{y} Verm.</th><th class='zk-num zk-sub'>{y} Eig.</th>"
    zusatz_detail_header += "<th class='zk-num'>Gesamt</th><th class='zk-num zk-sub'>Vermittler</th><th class='zk-num zk-sub'>Eigent\u00fcmer</th></tr>"

    zusatz_detail_rows = ""
    for z in zusatz_sorted:
        zusatz_detail_rows += f"<tr><td>{z['name']}</td><td class='zk-num'>{z['count']}</td>"
        for y in years_sorted:
            zy = z["per_year"][y]
            zusatz_detail_rows += f"<td class='zk-num'>{format_euro(zy['gesamt'])}</td>"
            zusatz_detail_rows += f"<td class='zk-num zk-sub'>{format_euro(zy['vermittler'])}</td>"
            zusatz_detail_rows += f"<td class='zk-num zk-sub'>{format_euro(zy['eigentuemer'])}</td>"
        zusatz_detail_rows += f"<td class='zk-num'><strong>{format_euro(z['total'])}</strong></td>"
        zusatz_detail_rows += f"<td class='zk-num zk-sub'>{format_euro(z['vermittler'])}</td>"
        zusatz_detail_rows += f"<td class='zk-num zk-sub'>{format_euro(z['eigentuemer'])}</td></tr>"

    # Totals row
    zusatz_detail_rows += "<tr class='zk-total'><td><strong>SUMME</strong></td><td></td>"
    for y in years_sorted:
        zt = zusatz_year_totals[y]
        zusatz_detail_rows += f"<td class='zk-num'><strong>{format_euro(zt['gesamt'])}</strong></td>"
        zusatz_detail_rows += f"<td class='zk-num zk-sub'><strong>{format_euro(zt['vermittler'])}</strong></td>"
        zusatz_detail_rows += f"<td class='zk-num zk-sub'><strong>{format_euro(zt['eigentuemer'])}</strong></td>"
    grand_v = sum(zt["vermittler"] for zt in zusatz_year_totals.values())
    grand_e = sum(zt["eigentuemer"] for zt in zusatz_year_totals.values())
    zusatz_detail_rows += f"<td class='zk-num'><strong>{format_euro(grand_v + grand_e)}</strong></td>"
    zusatz_detail_rows += f"<td class='zk-num zk-sub'><strong>{format_euro(grand_v)}</strong></td>"
    zusatz_detail_rows += f"<td class='zk-num zk-sub'><strong>{format_euro(grand_e)}</strong></td></tr>"

    zusatz_detail_table = f"""<div style="overflow-x:auto;"><table class="zusatz-detail-table">
        <thead>{zusatz_detail_header}</thead>
        <tbody>{zusatz_detail_rows}</tbody>
    </table></div>"""

    # Chart data: top 10 Zusatzkosten by absolute total (bar chart)
    top10_zusatz = zusatz_sorted[:10]
    zusatz_chart_labels = json.dumps([z["name"] for z in top10_zusatz])
    zusatz_chart_vermittler = json.dumps([round(z["vermittler"], 2) for z in top10_zusatz])
    zusatz_chart_eigentuemer = json.dumps([round(z["eigentuemer"], 2) for z in top10_zusatz])

    # Stacked bar chart per year for top 5 categories
    top5_zusatz = zusatz_sorted[:5]
    zusatz_yearly_labels = json.dumps([z["name"] for z in top5_zusatz])
    zusatz_yearly_datasets = []
    for i, y in enumerate(years_sorted):
        color = colors[i % len(colors)]
        zusatz_yearly_datasets.append({
            "label": str(y),
            "data": [round(z["per_year"][y]["gesamt"], 2) for z in top5_zusatz],
            "backgroundColor": color,
            "borderRadius": 4,
        })
    json_zusatz_yearly_datasets = json.dumps(zusatz_yearly_datasets)

    # --- Build Provisionen tab HTML ---
    prov_tab_parts = []
    for y in years:
        is_current = (y == current_year)
        collapsed = "none" if not is_current else "block"
        arrow = "\u25bc" if is_current else "\u25b6"
        section_id = f"prov-section-{y}"
        arrow_id = f"prov-arrow-{y}"

        # Provision KPIs f\u00fcr dieses Jahr
        k = kpis[y]
        mv_total = k["miete_vermittler"]
        mg_total = k["miete_gesamt"]
        prov_rate_avg = round(mv_total / mg_total * 100, 1) if mg_total > 0 else 0
        zk_verm_year = zusatz_year_totals[y]["vermittler"]

        # Zeilen pro Unterkunft (mit Eigent\u00fcmer)
        year_rows = []
        sum_mg = sum_mv = sum_zv = sum_pg = 0
        for pname, pprov in sorted(provision_by_prop.items()):
            if y not in pprov["years"]:
                continue
            py = pprov["years"][y]
            year_rows.append((pname, pprov["ort"], pprov.get("eigentuemer", ""), pprov.get("provision_pct_excel", ""), py))
            sum_mg += py["miete_gesamt"]
            sum_mv += py["miete_vermittler"]
            sum_zv += py["zusatz_vermittler"]
            sum_pg += py["provision_gesamt"]
        year_rows.sort(key=lambda x: -x[4]["provision_gesamt"])

        table_rows_html = ""
        sum_detail_plv = 0.0
        for pname, ort, eg, pct_excel, py in year_rows:
            plv = py.get("provision_laut_vertrag")
            plv_str = format_euro(plv) if plv is not None else "–"
            if plv: sum_detail_plv += plv
            table_rows_html += f'''
                    <tr>
                        <td>{pname}</td>
                        <td>{ort}</td>
                        <td style="font-size:12px;color:#555;">{eg}</td>
                        <td class="num" style="color:#0066cc;font-weight:600;">{pct_excel}</td>
                        <td class="num">{py["buchungen"]}</td>
                        <td class="num">{format_euro(py["miete_gesamt"])}</td>
                        <td class="num" style="color:#28a745;font-weight:600;">{plv_str}</td>
                    </tr>'''

        # Eigent\u00fcmer-Aggregation f\u00fcr dieses Jahr
        eg_rows = []
        for eg_name, eg_years in provision_by_eigentuemer.items():
            if y not in eg_years:
                continue
            ey = eg_years[y]
            eg_rows.append((eg_name, ey))
        eg_rows.sort(key=lambda x: -x[1]["provision_gesamt"])

        eg_table_html = ""
        sum_plv = sum(ey.get("provision_laut_vertrag", 0) for _, ey in eg_rows)
        for eg_name, ey in eg_rows:
            saetze = ey.get("provision_saetze", [])
            saetze_str = " / ".join(saetze) if saetze else "–"
            plv = ey.get("provision_laut_vertrag", 0)
            plv_str = format_euro(plv) if plv else "–"
            eg_table_html += f'''
                    <tr>
                        <td>{eg_name}</td>
                        <td class="num" style="color:#0066cc;font-weight:600;">{saetze_str}</td>
                        <td class="num">{ey["buchungen"]}</td>
                        <td class="num">{format_euro(ey["miete_gesamt"])}</td>
                        <td class="num" style="color:#28a745;font-weight:600;">{plv_str}</td>
                    </tr>'''

        sum_pct = round(sum_mv / sum_mg * 100, 1) if sum_mg > 0 else 0
        header_bg = "#0066cc" if is_current else "#f0f4fa"
        header_color = "#fff" if is_current else "#0066cc"

        prov_tab_parts.append(f'''
        <div style="margin-bottom:12px;border:1px solid #e0e0e0;border-radius:8px;overflow:hidden;">
            <div onclick="var c=document.getElementById('{section_id}');var a=document.getElementById('{arrow_id}');c.style.display=c.style.display==='none'?'block':'none';a.textContent=c.style.display==='none'?'\u25b6':'\u25bc';"
                 style="cursor:pointer;background:{header_bg};padding:12px 18px;display:flex;align-items:center;justify-content:space-between;user-select:none;">
                <span style="font-size:17px;font-weight:700;color:{header_color};">{y}</span>
                <div style="display:flex;align-items:center;gap:24px;">
                    <span style="font-size:13px;color:{header_color};opacity:0.9;">Provision laut Vertrag: {format_euro(sum_plv)}</span>
                    <span id="{arrow_id}" style="font-size:14px;color:{header_color};">{arrow}</span>
                </div>
            </div>
            <div id="{section_id}" style="display:{collapsed};padding:16px;">
                <div class="prop-detail-grid" style="margin-bottom:16px;">
                    <div class="prop-kpi"><div class="pk-label">Provision laut Vertrag</div><div class="pk-value green">{format_euro(sum_plv)}</div></div>
                    <div class="prop-kpi"><div class="pk-label">Miete gesamt</div><div class="pk-value">{format_euro(mg_total)}</div></div>
                    <div class="prop-kpi"><div class="pk-label">Provision aus Contao (Info)</div><div class="pk-value" style="color:#888;">{format_euro(mv_total)}</div></div>
                    <div class="prop-kpi"><div class="pk-label">Zusatzk. Vermittler (Info)</div><div class="pk-value" style="color:#888">{format_euro(zk_verm_year)}</div></div>
                </div>

                <!-- Eigent\u00fcmer-Aggregation -->
                <details open style="margin-bottom:20px;border:1px solid #d0e4ff;border-radius:6px;overflow:hidden;">
                    <summary style="cursor:pointer;background:#e8f0fe;padding:10px 14px;font-weight:600;color:#0066cc;list-style:none;display:flex;align-items:center;gap:8px;">
                        <span>&#9654;</span> Provision nach Eigent\u00fcmer ({len(eg_rows)} Eigent\u00fcmer)
                    </summary>
                    <div style="padding:12px;overflow-x:auto;">
                    <table class="prov-table">
                        <thead><tr>
                            <th>Eigent\u00fcmer</th>
                            <th class="num">Satz (Vertrag)</th>
                            <th class="num">Buchungen</th>
                            <th class="num">Miete gesamt</th>
                            <th class="num">Provision laut Vertrag</th>
                        </tr></thead>
                        <tbody>
                            {eg_table_html}
                            <tr class="prov-total">
                                <td><strong>SUMME {y}</strong></td>
                                <td></td>
                                <td></td>
                                <td class="num"><strong>{format_euro(sum_mg)}</strong></td>
                                <td class="num"><strong>{format_euro(sum_plv)}</strong></td>
                            </tr>
                        </tbody>
                    </table>
                    </div>
                </details>

                <!-- Details pro Unterkunft -->
                <details style="border:1px solid #e0e0e0;border-radius:6px;overflow:hidden;">
                    <summary style="cursor:pointer;background:#f8f9fa;padding:10px 14px;font-weight:600;color:#444;list-style:none;display:flex;align-items:center;gap:8px;">
                        <span>&#9654;</span> Details pro Unterkunft ({len(year_rows)} Unterk\u00fcnfte)
                    </summary>
                    <div style="padding:12px;overflow-x:auto;">
                    <table class="prov-table">
                        <thead><tr>
                            <th>Unterkunft</th><th>Ort</th>
                            <th>Eigent\u00fcmer</th>
                            <th class="num">Satz (Vertrag)</th>
                            <th class="num">Buchungen</th>
                            <th class="num">Miete gesamt</th>
                            <th class="num">Provision laut Vertrag</th>
                        </tr></thead>
                        <tbody>
                            {table_rows_html}
                            <tr class="prov-total">
                                <td colspan="3"><strong>SUMME {y}</strong></td>
                                <td></td>
                                <td></td>
                                <td class="num"><strong>{format_euro(sum_mg)}</strong></td>
                                <td class="num"><strong>{format_euro(sum_detail_plv)}</strong></td>
                            </tr>
                        </tbody>
                    </table>
                    </div>
                </details>
            </div>
        </div>''')

    # --- Build Provisionen archive (2017-2023) as collapsible ---
    prov_archive_parts = []
    for y in years_archive:
        k = kpis[y]
        mv_total = k["miete_vermittler"]
        mg_total = k["miete_gesamt"]
        prov_rate_avg = round(mv_total / mg_total * 100, 1) if mg_total > 0 else 0
        zk_verm_year = zusatz_year_totals[y]["vermittler"]
        prov_archive_parts.append(f'''
        <div class="prop-section">
            <h4 style="color:#888;border-bottom:1px solid #e0e0e0;padding-bottom:6px;font-size:16px;">{y}</h4>
            <div class="prop-detail-grid" style="margin-bottom:12px;">
                <div class="prop-kpi"><div class="pk-label">Provision (Miete Verm.)</div><div class="pk-value green">{format_euro(mv_total)}</div></div>
                <div class="prop-kpi"><div class="pk-label">Miete gesamt</div><div class="pk-value">{format_euro(mg_total)}</div></div>
                <div class="prop-kpi"><div class="pk-label">Ø Provisionssatz</div><div class="pk-value">{format_german_number(prov_rate_avg, 1)} %</div></div>
                <div class="prop-kpi"><div class="pk-label">Zusatzk. Vermittler (Info)</div><div class="pk-value" style="color:#888">{format_euro(zk_verm_year)}</div></div>
            </div>
        </div>''')
    archive_html = ""
    if prov_archive_parts:
        archive_html = f'''
    <details style="margin-top:24px;">
        <summary style="cursor:pointer;font-weight:600;color:#888;padding:10px 0;border-top:2px solid #e0e0e0;list-style:none;display:flex;align-items:center;gap:8px;">
            <span style="font-size:18px;">&#9654;</span> Archiv 2017–2023
        </summary>
        {"".join(prov_archive_parts)}
    </details>'''
    prov_tab_html = "\n".join(prov_tab_parts) + archive_html

    # --- Build Reiseprofile tab HTML ---
    # Stacked bar chart data: profiles per year
    profile_chart_datasets = []
    for p in profile_order:
        vals = [profiles_by_year[y].get(p, 0) for y in years]
        if sum(vals) == 0:
            continue
        profile_chart_datasets.append({
            "label": p,
            "data": vals,
            "backgroundColor": profile_colors.get(p, "#ccc"),
            "borderRadius": 4,
        })
    json_profile_datasets = json.dumps(profile_chart_datasets)

    # Property table: sorted by dog percentage (most interesting for targeting)
    profile_table_rows = ""
    prop_profile_list = []
    for pname, pd in prop_profiles.items():
        if pd["buchungen"] < 5:
            continue
        total_b = pd["buchungen"]
        hund_b = pd["total"].get("Urlaub mit Hund", 0)
        familie_b = pd["total"].get("Familie mit Kleinkind", 0)
        gruppe_b = pd["total"].get("Größere Reisegruppe", 0)
        wellness_b = pd["total"].get("Wellness-Gäste", 0)
        eauto_b = pd["total"].get("E-Auto Reisende", 0)
        paar_b = pd["total"].get("Paare/Einzelreisende", 0)
        prop_profile_list.append({
            "name": pname,
            "total": total_b,
            "hund": hund_b, "hund_pct": round(hund_b / total_b * 100, 1),
            "familie": familie_b, "familie_pct": round(familie_b / total_b * 100, 1),
            "gruppe": gruppe_b, "gruppe_pct": round(gruppe_b / total_b * 100, 1),
            "wellness": wellness_b, "wellness_pct": round(wellness_b / total_b * 100, 1),
            "eauto": eauto_b, "eauto_pct": round(eauto_b / total_b * 100, 1),
            "paar": paar_b, "paar_pct": round(paar_b / total_b * 100, 1),
        })
    prop_profile_list.sort(key=lambda x: -x["hund_pct"])

    for pp in prop_profile_list:
        # Visual bar for profile distribution
        bar_html = ""
        for label, key, color in [
            ("Paar", "paar", "#0066cc"), ("Hund", "hund", "#ff6b6b"),
            ("Gruppe", "gruppe", "#ffa500"), ("Familie", "familie", "#4ecdc4"),
            ("Wellness", "wellness", "#aa96da"), ("E-Auto", "eauto", "#95e1d3")
        ]:
            pct = pp[f"{key}_pct"]
            if pct > 0:
                bar_html += f'<div style="width:{max(pct, 2)}%;background:{color};height:100%;display:inline-block;" title="{label}: {pp[key]} ({pct}%)"></div>'

        profile_table_rows += f'''<tr>
            <td>{pp["name"]}</td>
            <td class="num">{pp["total"]}</td>
            <td class="num" style="color:#ff6b6b;font-weight:600;">{pp["hund"]} <small>({pp["hund_pct"]}%)</small></td>
            <td class="num" style="color:#4ecdc4;">{pp["familie"]} <small>({pp["familie_pct"]}%)</small></td>
            <td class="num" style="color:#ffa500;">{pp["gruppe"]} <small>({pp["gruppe_pct"]}%)</small></td>
            <td class="num" style="color:#aa96da;">{pp["wellness"]} <small>({pp["wellness_pct"]}%)</small></td>
            <td class="num" style="color:#0066cc;">{pp["paar"]} <small>({pp["paar_pct"]}%)</small></td>
            <td style="min-width:200px;"><div style="display:flex;height:16px;border-radius:4px;overflow:hidden;">{bar_html}</div></td>
        </tr>'''

    # Top Hund properties chart (top 15)
    top_hund = sorted(prop_profile_list, key=lambda x: -x["hund_pct"])[:15]
    json_hund_labels = json.dumps([p["name"] for p in top_hund])
    json_hund_pcts = json.dumps([p["hund_pct"] for p in top_hund])
    json_hund_counts = json.dumps([p["hund"] for p in top_hund])

    # --- Build property data JSON for embedded JS ---
    prop_json_data = {}
    for pname, pdata in property_data.items():
        pj = {"ort": pdata["ort"], "years": {}}
        for y, yd in pdata["years"].items():
            pj["years"][y] = {
                "buchungen": yd["buchungen"],
                "naechte": yd["naechte"],
                "reisepreis": round(yd["reisepreis"], 2),
                "miete_gesamt": round(yd["miete_gesamt"], 2),
                "miete_vermittler": round(yd["miete_vermittler"], 2),
                "miete_eigentuemer": round(yd["miete_eigentuemer"], 2),
                "provision_pct": yd["provision_pct"],
                "avg_preis_nacht": yd["avg_preis_nacht"],
                "avg_aufenthalt": yd["avg_aufenthalt"],
                "belegung_pct": yd["belegung_pct"],
                "channels": [list(c) for c in yd["channels"]] if isinstance(yd["channels"], list) and yd["channels"] and isinstance(yd["channels"][0], tuple) else yd["channels"],
                "zusatzkosten": yd["zusatzkosten"]
            }
        prop_json_data[pname] = pj
    json_property_data = json.dumps(prop_json_data, ensure_ascii=False)

    property_options = "\n".join(
        f'<option value="{pname}">{pname} ({pdata["ort"]})</option>'
        for pname, pdata in sorted(property_data.items())
    )

    # --- JSON data for charts ---
    json_years = json.dumps(years)
    json_month_names = json.dumps(month_names)
    json_colors = json.dumps(colors)

    # Monthly line chart datasets
    line_datasets = []
    for i, y in enumerate(years):
        color = colors[i % len(colors)]
        line_datasets.append({
            "label": str(y),
            "data": [round(v, 1) for v in monthly_data[y]],
            "borderColor": color,
            "backgroundColor": color + "33",
            "tension": 0.3,
            "fill": False,
            "borderWidth": 2,
            "pointRadius": 4,
        })
    json_line_datasets = json.dumps(line_datasets)

    # Top properties bar chart
    prop_labels = json.dumps([p[0] for p in top_props])
    prop_values = json.dumps([p[1] for p in top_props])

    # Channels doughnut (legacy, kept for overall chart)
    channel_labels = json.dumps([c[0] for c in channels])
    channel_values = json.dumps([c[1] for c in channels])
    channel_colors = json.dumps([colors[i % len(colors)] for i in range(len(channels))])

    # --- Build Vertriebskanäle per-year HTML ---
    channel_year_html_parts = []
    channel_chart_js_parts = []
    for idx, y in enumerate(years_sorted):
        cy = channels_by_year[y]
        total = cy["total"]
        chs = cy["channels"]

        # Table rows with expandable groups (Ostseeliebe + Portalbuchungen)
        portal_prov_year = cy.get("portal_provision", 0)
        table_rows = ""
        for gi, ch in enumerate(chs):
            grp_cls = f"ch-g{gi}-{y}"
            is_portal = ch["name"] == "Portalbuchungen"
            bg_color = "#f0f7ff" if not is_portal else "#fff7f0"
            sub_bg = "#f8fbff" if not is_portal else "#fffbf8"
            prov_cell = f'<td class="zk-num"><strong style="color:#cc3333">{format_euro(ch["provision"])}</strong></td>' if is_portal else '<td class="zk-num" style="color:#ccc">–</td>'
            table_rows += f'''<tr class="ch-group" onclick="this.parentElement.querySelectorAll('.{grp_cls}').forEach(function(r){{r.style.display=r.style.display==='none'?'table-row':'none'}});" style="cursor:pointer;background:{bg_color}">
                <td><strong>&#9654; {ch["name"]}</strong></td>
                <td class="zk-num"><strong>{format_german_number(ch["count"], 0)}</strong></td>
                <td class="zk-num"><strong>{ch["pct"]} %</strong></td>
                {prov_cell}
            </tr>'''
            for sub in ch["sub"]:
                sub_prov = sub.get("provision", 0)
                sub_prov_cell = f'<td class="zk-num" style="color:#cc3333;font-size:12px">{format_euro(sub_prov)}</td>' if is_portal else '<td class="zk-num" style="color:#ccc">–</td>'
                table_rows += f'''<tr class="{grp_cls}" style="display:none;background:{sub_bg}">
                    <td style="padding-left:28px">{sub["name"]}</td>
                    <td class="zk-num">{format_german_number(sub["count"], 0)}</td>
                    <td class="zk-num">{sub["pct"]} %</td>
                    {sub_prov_cell}
                </tr>'''
        table_rows += f'''<tr class="zk-total">
            <td><strong>Gesamt</strong></td>
            <td class="zk-num"><strong>{format_german_number(total, 0)}</strong></td>
            <td class="zk-num"><strong>100 %</strong></td>
            <td class="zk-num"><strong style="color:#cc3333">{format_euro(portal_prov_year)}</strong></td>
        </tr>'''

        # Chart data for this year (Ostseeliebe vs Portalbuchungen)
        chart_id = f"channelChart{y}"
        ch_labels = json.dumps([ch["name"] for ch in chs])
        ch_values = json.dumps([ch["count"] for ch in chs])
        ch_colors = json.dumps(["#0066cc", "#ff6b6b"][:len(chs)])

        channel_year_html_parts.append(f'''
        <div class="year-section">
            <h3 class="year-title">{y}</h3>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;align-items:start">
                <div class="zusatz-summary">
                    <table class="zusatz-table">
                        <thead><tr><th>Vertriebskanal</th><th class="zk-num">Buchungen</th><th class="zk-num">Anteil</th><th class="zk-num">Portalprovision</th></tr></thead>
                        <tbody>{table_rows}</tbody>
                    </table>
                    <div style="font-size:11px;color:#888;margin-top:6px">&#9654; Klicke auf eine Gruppe um die Einzelkanäle aufzuklappen</div>
                </div>
                <div class="chart-wrapper doughnut-chart" style="max-width:360px;margin:0 auto">
                    <canvas id="{chart_id}"></canvas>
                </div>
            </div>
        </div>''')

        channel_chart_js_parts.append(f'''
        new Chart(document.getElementById('{chart_id}'), {{
            type: 'doughnut',
            data: {{
                labels: {ch_labels},
                datasets: [{{
                    data: {ch_values},
                    backgroundColor: {ch_colors},
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }},
                    tooltip: {{ callbacks: {{ label: function(c) {{
                        var total = c.dataset.data.reduce(function(a,b){{return a+b}},0);
                        var pct = (c.raw/total*100).toFixed(1);
                        return c.label + ': ' + c.raw.toLocaleString('de-DE') + ' (' + pct + ' %)';
                    }} }} }}
                }}
            }}
        }});''')

    channel_year_html = "\n".join(channel_year_html_parts)
    channel_chart_js = "\n".join(channel_chart_js_parts)

    # Locations bar
    ort_labels = json.dumps([o["ort"] for o in orte])
    ort_values = json.dumps([o["total_buchungen"] for o in orte])
    ort_colors = json.dumps([colors[i % len(colors)] for i in range(len(orte))])


    # ---------------------------------------------------------------------------
    # Tab layout — THIS is the single place to define and order tabs.
    # To add a new tab:
    #   1. Add an entry here: ("your_tab_id", "Tab Label")
    #   2. Add the matching content block in tab_contents below
    # The tab navigation and active-state logic are generated automatically.
    # ---------------------------------------------------------------------------

    # --- Build Apartmenth\u00e4user tab HTML ---
    # Embed all years' data as JSON so the client can filter by year
    ah_json_years = json.dumps(years_sorted)
    # Build per-house data structure for JS
    ah_js_data = {}
    for haus_name, year_units in haus_year_data.items():
        ah_js_data[haus_name] = {}
        for y, units in year_units.items():
            ah_js_data[haus_name][y] = [
                {"name": uname, "buchungen": v["buchungen"], "umsatz": round(v["umsatz"], 2), "miete_eigentuemer": round(v["miete_eigentuemer"], 2)}
                for uname, v in units.items()
            ]
    ah_json_data = json.dumps(ah_js_data, ensure_ascii=False)

    # Static card HTML (all years combined, for initial render / "Alle Jahre")
    def _build_ah_cards(year_filter=None):
        cards = []
        for haus_name, year_units in sorted(haus_year_data.items()):
            unit_totals = defaultdict(lambda: {"buchungen": 0, "umsatz": 0.0, "miete_eigentuemer": 0.0})
            for y, units in year_units.items():
                if year_filter and y != year_filter:
                    continue
                for uname, v in units.items():
                    unit_totals[uname]["buchungen"] += v["buchungen"]
                    unit_totals[uname]["umsatz"] += v["umsatz"]
                    unit_totals[uname]["miete_eigentuemer"] += v["miete_eigentuemer"]
            gb = sum(v["buchungen"] for v in unit_totals.values())
            gu = sum(v["umsatz"] for v in unit_totals.values())
            gme = sum(v["miete_eigentuemer"] for v in unit_totals.values())
            if gb == 0:
                continue
            obj_count = len(unit_totals)
            max_b = max((v["buchungen"] for v in unit_totals.values()), default=1) or 1
            rows = ""
            for uname, v in sorted(unit_totals.items(), key=lambda x: -x[1]["buchungen"]):
                bar_pct = round(v["buchungen"] / max_b * 100)
                umsatz_str = format_euro(v["umsatz"]) if v["umsatz"] > 0 else "\u2013"
                me_str = format_euro(v["miete_eigentuemer"]) if v["miete_eigentuemer"] > 0 else "\u2013"
                rows += f'''<tr>
                        <td class="ah-obj-name">{uname}</td>
                        <td class="ah-obj-bar"><div class="ah-bar-wrap"><div class="ah-bar" style="width:{bar_pct}%"></div><span class="ah-bar-label">{v["buchungen"]}</span></div></td>
                        <td class="ah-obj-umsatz">{umsatz_str}</td>
                        <td class="ah-obj-me">{me_str}</td>
                    </tr>'''
            # Totals row
            rows += f'''<tr class="ah-total-row">
                        <td class="ah-obj-name"><strong>Gesamt</strong></td>
                        <td class="ah-obj-bar"><strong>{gb} Buchungen</strong></td>
                        <td class="ah-obj-umsatz"><strong>{format_euro(gu)}</strong></td>
                        <td class="ah-obj-me"><strong>{format_euro(gme)}</strong></td>
                    </tr>'''
            suffix = "en" if obj_count != 1 else ""
            cards.append((gb, gu, gme, haus_name, f'''<div class="ah-card" data-buchungen="{gb}" data-umsatz="{round(gu,2)}" data-me="{round(gme,2)}" data-name="{haus_name}">
                    <div class="ah-header">
                        <span class="ah-name">{haus_name}</span>
                        <span class="ah-kpi-row">
                            <span class="ah-kpi-b">{gb} Buchungen</span>
                            <span class="ah-kpi-u">{format_euro(gu)}</span>
                            <span class="ah-kpi-me">Eig.: {format_euro(gme)}</span>
                            <span class="ah-kpi-n">{obj_count} Unterkunft{suffix}</span>
                        </span>
                    </div>
                    <table class="ah-obj-table">
                        <thead><tr>
                            <th class="ah-th-name">Unterkunft</th>
                            <th class="ah-th-bar">Buchungen</th>
                            <th class="ah-th-num">Reisepreis</th>
                            <th class="ah-th-num">Miete Eig.</th>
                        </tr></thead>
                        <tbody>{rows}</tbody>
                    </table>
                </div>'''))
        cards.sort(key=lambda x: -x[0])
        return "\n".join(c[4] for c in cards)

    ah_cards_html = _build_ah_cards()

    # Year-filter buttons
    ah_year_btns = "\n".join(
        f'<button class="ah-yr-btn" data-year="{y}" onclick="ahFilter({y},this)">{y}</button>'
        for y in years_sorted
    )

    apartmenthaus_tab_html = f'''<div class="chart-container">
        <h3>Apartmenth\u00e4user \u2013 Gesamtsummen &amp; Buchungsvergleich</h3>
        <p style="color:var(--color-text-muted);font-size:13px;margin-bottom:14px;">Reisepreis und Miete Eigent\u00fcmer je Apartmenthaus, aufgeschl\u00fcsselt nach Unterkunft.</p>
        <div class="ah-filter-row">
            {ah_year_btns}
            <button class="ah-yr-btn ah-yr-all active" onclick="ahFilter(null,this)">Alle Jahre</button>
            <span style="margin-left:auto;font-size:13px;color:var(--color-text-muted);">Sortieren:
            <select id="ahSort" onchange="ahSort(this.value)" style="border:1px solid var(--color-border);border-radius:6px;padding:3px 8px;font-size:13px;">
                <option value="buchungen">Buchungen \u2193</option>
                <option value="umsatz">Reisepreis \u2193</option>
                <option value="me">Miete Eig. \u2193</option>
                <option value="name">Name A\u2013Z</option>
            </select></span>
        </div>
        <div class="ah-kpi-summary" id="ahKpiRow">
            <div class="ah-summary-kpi"><div class="ah-summary-val" id="ahKpiHaeuser">\u2013</div><div class="ah-summary-lbl">H\u00e4user</div></div>
            <div class="ah-summary-kpi"><div class="ah-summary-val" id="ahKpiBuchungen">\u2013</div><div class="ah-summary-lbl">Buchungen</div></div>
            <div class="ah-summary-kpi"><div class="ah-summary-val" id="ahKpiUmsatz">\u2013</div><div class="ah-summary-lbl">Reisepreis gesamt</div></div>
            <div class="ah-summary-kpi"><div class="ah-summary-val" id="ahKpiMe" style="color:#065f46;">\u2013</div><div class="ah-summary-lbl">Miete Eigent\u00fcmer gesamt</div></div>
        </div>

        <!-- Gesamt\u00fcbersicht-Tabelle -->
        <div class="ah-overview-wrap">
            <h4 style="font-size:13px;font-weight:600;color:var(--color-text-muted);margin-bottom:8px;text-transform:uppercase;letter-spacing:.5px;">Gesamt\u00fcbersicht alle H\u00e4user</h4>
            <table class="ah-overview-table" id="ahOverviewTable">
                <thead><tr>
                    <th>Haus</th>
                    <th class="ah-ov-num">Buchungen</th>
                    <th class="ah-ov-num">Reisepreis</th>
                    <th class="ah-ov-num">Miete Eig.</th>
                    <th class="ah-ov-num">Eig.-Anteil</th>
                </tr></thead>
                <tbody id="ahOverviewBody"></tbody>
                <tfoot id="ahOverviewFoot"></tfoot>
            </table>
        </div>

        <!-- Detailkarten -->
        <h4 style="font-size:13px;font-weight:600;color:var(--color-text-muted);margin:18px 0 10px;text-transform:uppercase;letter-spacing:.5px;">Detailansicht je Haus</h4>
        <div class="ah-grid" id="ahGrid">
            {ah_cards_html}
        </div>
    </div>
    <style>
        .ah-filter-row {{ display:flex; gap:6px; flex-wrap:wrap; align-items:center; margin-bottom:14px; }}
        .ah-yr-btn {{ padding:4px 13px; border:1px solid var(--color-border); border-radius:16px; background:#fff; cursor:pointer; font-size:12px; color:#555; transition:all .15s; }}
        .ah-yr-btn.active, .ah-yr-btn:hover {{ background:var(--color-primary); color:#fff; border-color:var(--color-primary); }}
        .ah-kpi-summary {{ display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin-bottom:16px; }}
        .ah-summary-kpi {{ background:var(--color-bg); border-radius:var(--radius-sm); padding:10px; text-align:center; }}
        .ah-summary-val {{ font-size:20px; font-weight:700; color:var(--color-primary); }}
        .ah-summary-lbl {{ font-size:11px; color:var(--color-text-muted); margin-top:2px; }}
        /* Overview table */
        .ah-overview-wrap {{ background:#fff; border-radius:var(--radius-sm); padding:14px; box-shadow:var(--shadow-card); margin-bottom:18px; overflow-x:auto; }}
        .ah-overview-table {{ width:100%; border-collapse:collapse; font-size:13px; }}
        .ah-overview-table th {{ background:var(--color-bg); color:var(--color-text-muted); font-size:11px; text-transform:uppercase; letter-spacing:.4px; padding:6px 10px; text-align:left; }}
        .ah-ov-num {{ text-align:right !important; font-variant-numeric:tabular-nums; }}
        .ah-overview-table td {{ padding:6px 10px; border-bottom:1px solid var(--color-border); }}
        .ah-overview-table tbody tr:hover {{ background:#f8fafc; }}
        .ah-overview-table tfoot td {{ padding:8px 10px; border-top:2px solid var(--color-primary); font-weight:700; background:var(--color-bg); }}
        .ah-ov-me {{ color:#065f46; font-weight:600; }}
        .ah-ov-pct {{ color:#888; font-size:12px; }}
        /* Cards */
        .ah-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(420px,1fr)); gap:14px; }}
        .ah-card {{ background:#fff; border-radius:var(--radius-sm); padding:14px; box-shadow:var(--shadow-card); }}
        .ah-header {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:10px; gap:8px; }}
        .ah-name {{ font-weight:600; font-size:13px; color:var(--color-text); flex:1; }}
        .ah-kpi-row {{ display:flex; flex-wrap:wrap; gap:4px; justify-content:flex-end; }}
        .ah-kpi-b {{ background:#dbeafe; color:#1e40af; padding:2px 7px; border-radius:10px; font-size:11px; font-weight:600; }}
        .ah-kpi-u {{ background:#fef9c3; color:#854d0e; padding:2px 7px; border-radius:10px; font-size:11px; font-weight:600; }}
        .ah-kpi-me {{ background:#d1fae5; color:#065f46; padding:2px 7px; border-radius:10px; font-size:11px; font-weight:600; }}
        .ah-kpi-n {{ background:var(--color-bg); color:#555; padding:2px 7px; border-radius:10px; font-size:11px; }}
        .ah-obj-table {{ width:100%; border-collapse:collapse; }}
        .ah-th-name,.ah-th-bar,.ah-th-num {{ font-size:10px; color:var(--color-text-muted); text-transform:uppercase; letter-spacing:.4px; padding:4px 3px; border-bottom:1px solid var(--color-border); }}
        .ah-th-num {{ text-align:right; }}
        .ah-obj-table tbody tr:not(.ah-total-row):not(:last-child) td {{ border-bottom:1px solid var(--color-border); }}
        .ah-total-row td {{ border-top:2px solid var(--color-primary) !important; background:var(--color-bg); padding:5px 3px; }}
        .ah-obj-table td {{ padding:4px 3px; vertical-align:middle; }}
        .ah-obj-name {{ font-size:12px; color:var(--color-text); width:28%; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
        .ah-obj-bar {{ width:36%; padding-right:6px; }}
        .ah-obj-umsatz {{ font-size:11px; color:#854d0e; text-align:right; white-space:nowrap; width:18%; }}
        .ah-obj-me {{ font-size:11px; color:#065f46; text-align:right; white-space:nowrap; width:18%; font-weight:500; }}
        .ah-bar-wrap {{ display:flex; align-items:center; gap:5px; }}
        .ah-bar {{ height:11px; background:var(--color-primary); border-radius:3px; min-width:3px; }}
        .ah-bar-label {{ font-size:11px; color:var(--color-text); white-space:nowrap; }}
        @media(max-width:700px) {{ .ah-kpi-summary {{ grid-template-columns:repeat(2,1fr); }} .ah-grid {{ grid-template-columns:1fr; }} }}
    </style>
    <script>
    (function() {{
        const AH_DATA = {ah_json_data};
        const AH_YEARS = {ah_json_years};
        let currentYear = null;

        function fmtEur(v) {{
            return v.toLocaleString('de-DE',{{minimumFractionDigits:0,maximumFractionDigits:0}}) + '\u202f\u20ac';
        }}
        function pct(a, b) {{
            if (!b) return '\u2013';
            return Math.round(a / b * 100) + '\u202f%';
        }}

        function updateCards(year) {{
            currentYear = year;
            const grid = document.getElementById('ahGrid');
            const sortVal = document.getElementById('ahSort').value;
            let cards = [];
            for (const [hausName, yearData] of Object.entries(AH_DATA)) {{
                let units = {{}};
                const yrs = year ? [String(year)] : AH_YEARS.map(String);
                yrs.forEach(y => {{
                    (yearData[y] || []).forEach(u => {{
                        if (!units[u.name]) units[u.name] = {{buchungen:0, umsatz:0, miete_eigentuemer:0}};
                        units[u.name].buchungen += u.buchungen;
                        units[u.name].umsatz += u.umsatz;
                        units[u.name].miete_eigentuemer += (u.miete_eigentuemer || 0);
                    }});
                }});
                const gb = Object.values(units).reduce((s,u) => s + u.buchungen, 0);
                const gu = Object.values(units).reduce((s,u) => s + u.umsatz, 0);
                const gme = Object.values(units).reduce((s,u) => s + u.miete_eigentuemer, 0);
                if (gb === 0) continue;
                cards.push({{hausName, units, gb, gu, gme}});
            }}
            // Sort
            if (sortVal === 'buchungen') cards.sort((a,b) => b.gb - a.gb);
            else if (sortVal === 'umsatz') cards.sort((a,b) => b.gu - a.gu);
            else if (sortVal === 'me') cards.sort((a,b) => b.gme - a.gme);
            else cards.sort((a,b) => a.hausName.localeCompare(b.hausName, 'de'));

            // --- Gesamt\u00fcbersicht-Tabelle ---
            const tbody = document.getElementById('ahOverviewBody');
            const tfoot = document.getElementById('ahOverviewFoot');
            let totB=0, totU=0, totMe=0;
            tbody.innerHTML = cards.map(c => {{
                totB += c.gb; totU += c.gu; totMe += c.gme;
                const p = pct(c.gme, c.gu);
                return `<tr>
                    <td>${{c.hausName}}</td>
                    <td class="ah-ov-num">${{c.gb.toLocaleString('de-DE')}}</td>
                    <td class="ah-ov-num">${{fmtEur(c.gu)}}</td>
                    <td class="ah-ov-num ah-ov-me">${{fmtEur(c.gme)}}</td>
                    <td class="ah-ov-num ah-ov-pct">${{p}}</td>
                </tr>`;
            }}).join('');
            tfoot.innerHTML = `<tr>
                <td><strong>Gesamt (${{cards.length}} H\u00e4user)</strong></td>
                <td class="ah-ov-num"><strong>${{totB.toLocaleString('de-DE')}}</strong></td>
                <td class="ah-ov-num"><strong>${{fmtEur(totU)}}</strong></td>
                <td class="ah-ov-num ah-ov-me"><strong>${{fmtEur(totMe)}}</strong></td>
                <td class="ah-ov-num ah-ov-pct"><strong>${{pct(totMe, totU)}}</strong></td>
            </tr>`;

            // --- Detailkarten ---
            grid.innerHTML = cards.map(c => {{
                const maxB = Math.max(...Object.values(c.units).map(u => u.buchungen), 1);
                const rowsSorted = Object.entries(c.units).sort((a,b) => b[1].buchungen - a[1].buchungen);
                const rows = rowsSorted.map(([uname, v]) => {{
                    const barPct = Math.round(v.buchungen / maxB * 100);
                    const uStr = v.umsatz > 0 ? fmtEur(v.umsatz) : '\u2013';
                    const meStr = v.miete_eigentuemer > 0 ? fmtEur(v.miete_eigentuemer) : '\u2013';
                    return `<tr>
                        <td class="ah-obj-name" title="${{uname}}">${{uname}}</td>
                        <td class="ah-obj-bar"><div class="ah-bar-wrap"><div class="ah-bar" style="width:${{barPct}}%"></div><span class="ah-bar-label">${{v.buchungen}}</span></div></td>
                        <td class="ah-obj-umsatz">${{uStr}}</td>
                        <td class="ah-obj-me">${{meStr}}</td>
                    </tr>`;
                }}).join('');
                const totalRow = `<tr class="ah-total-row">
                    <td class="ah-obj-name"><strong>Gesamt</strong></td>
                    <td class="ah-obj-bar"><strong>${{c.gb}} Buchungen</strong></td>
                    <td class="ah-obj-umsatz"><strong>${{fmtEur(c.gu)}}</strong></td>
                    <td class="ah-obj-me"><strong>${{fmtEur(c.gme)}}</strong></td>
                </tr>`;
                const n = Object.keys(c.units).length;
                return `<div class="ah-card">
                    <div class="ah-header">
                        <span class="ah-name">${{c.hausName}}</span>
                        <span class="ah-kpi-row">
                            <span class="ah-kpi-b">${{c.gb}} Buchungen</span>
                            <span class="ah-kpi-u">${{fmtEur(c.gu)}}</span>
                            <span class="ah-kpi-me">Eig.: ${{fmtEur(c.gme)}}</span>
                            <span class="ah-kpi-n">${{n}} Unterkunft${{n!==1?'en':''}}</span>
                        </span>
                    </div>
                    <table class="ah-obj-table">
                        <thead><tr>
                            <th class="ah-th-name">Unterkunft</th>
                            <th class="ah-th-bar">Buchungen</th>
                            <th class="ah-th-num">Reisepreis</th>
                            <th class="ah-th-num">Miete Eig.</th>
                        </tr></thead>
                        <tbody>${{rows}}${{totalRow}}</tbody>
                    </table>
                </div>`;
            }}).join('\\n');

            // Update KPI-Leiste
            document.getElementById('ahKpiHaeuser').textContent = cards.length;
            document.getElementById('ahKpiBuchungen').textContent = totB.toLocaleString('de-DE');
            document.getElementById('ahKpiUmsatz').textContent = fmtEur(totU);
            document.getElementById('ahKpiMe').textContent = fmtEur(totMe);
        }}

        window.ahFilter = function(year, btn) {{
            document.querySelectorAll('.ah-yr-btn').forEach(b => b.classList.remove('active'));
            if (btn) btn.classList.add('active');
            updateCards(year);
        }};
        window.ahSort = function(val) {{ updateCards(currentYear); }};

        // Initial render
        updateCards(null);
    }})();
    </script>'''

    # --- Build Umsatz-Ranking tab HTML ---
    ranking_data = []
    current_year = datetime.now().year
    for pname, pd in property_data.items():
        total_umsatz = sum(pd["years"].get(y, {}).get("reisepreis", 0) for y in years)
        if total_umsatz == 0:
            continue
        total_naechte = sum(pd["years"].get(y, {}).get("naechte", 0) for y in years)
        total_buchungen = sum(pd["years"].get(y, {}).get("buchungen", 0) for y in years)
        # Auslastung: gebuchte N\u00e4chte / verf\u00fcgbare N\u00e4chte im Zeitraum
        # Verf\u00fcgbare N\u00e4chte = Anzahl Jahre \u00d7 365 (ohne Schaltjahre-Korrektur)
        verfuegbar = len([y for y in years if y <= current_year]) * 365
        auslastung = round(total_naechte / verfuegbar * 100, 1) if verfuegbar > 0 else 0
        # J\u00e4hrliche N\u00e4chte und Auslastung
        per_year_naechte = {y: pd["years"].get(y, {}).get("naechte", 0) for y in years}
        per_year_asl = {
            y: round(pd["years"].get(y, {}).get("naechte", 0) / 365 * 100, 1)
            if y <= current_year else 0
            for y in years
        }
        ranking_data.append({
            "name": pname,
            "ort": pd["ort"],
            "total": total_umsatz,
            "per_year": {y: pd["years"].get(y, {}).get("reisepreis", 0) for y in years},
            "per_year_naechte": per_year_naechte,
            "per_year_asl": per_year_asl,
            "buchungen": total_buchungen,
            "naechte": total_naechte,
            "auslastung": auslastung,
        })
    # Sortierung ausschließlich nach aktuellem Jahr (höchster Umsatz zuerst)
    ranking_data.sort(key=lambda x: -x["per_year"].get(current_year, 0))

    max_umsatz = max((x["per_year"].get(current_year, 0) for x in ranking_data), default=1) or 1

    # Spaltenk\u00f6pfe: Gruppe pro Jahr (Umsatz | N\u00e4chte | Auslastung), dann Gesamt
    year_header_top = "".join(
        f'<th class="num" colspan="3" style="text-align:center;background:#0066cc;color:#fff;font-weight:700;font-size:13px;letter-spacing:0.5px">{y}</th>'
        for y in years
    )
    year_header_sub = "".join(
        '<th class="num" style="font-size:11px;color:#fff;font-weight:500;background:#1a7de0">Umsatz</th>'
        '<th class="num" style="font-size:11px;color:#fff;font-weight:500;background:#1a7de0">N\u00e4chte</th>'
        '<th class="num" style="font-size:11px;color:#fff;font-weight:500;background:#1a7de0">Ausl.</th>'
        for _ in years
    )

    ranking_rows = ""
    MEDALS = {1: "&#127949;", 2: "&#127950;", 3: "&#127951;"}
    for i, item in enumerate(ranking_data, 1):
        cur_umsatz = item["per_year"].get(current_year, 0)
        bar_pct = round(cur_umsatz / max_umsatz * 100) if cur_umsatz else 0
        medal = MEDALS.get(i, f"<span style='color:#999;font-size:12px'>#{i}</span>")
        _dash = "<span style=\"color:#ccc\">&#8211;</span>"
        year_cells = ""
        for y in years:
            umsatz_y = item["per_year"].get(y, 0)
            naechte_y = item["per_year_naechte"].get(y, 0)
            asl_y = item["per_year_asl"].get(y, 0)
            asl_col = "#2d7a2d" if asl_y >= 60 else ("#e67e00" if asl_y >= 30 else "#cc3333")
            # Aktuelles Jahr farblich hervorheben
            is_cur = (y == current_year)
            cell_style = "font-size:12px;background:#f0f7ff;font-weight:600" if is_cur else "font-size:12px"
            year_cells += (
                f'<td class="num" style="{cell_style}">' + (format_euro(umsatz_y) if umsatz_y > 0 else _dash) + '</td>'
                f'<td class="num" style="font-size:12px;color:#555{";background:#f0f7ff" if is_cur else ""}">' + (str(int(naechte_y)) if naechte_y > 0 else _dash) + '</td>'
                f'<td class="num" style="font-size:12px;color:{asl_col};font-weight:600{";background:#f0f7ff" if is_cur else ""}">' + (f'{asl_y:.1f}\u00a0%' if naechte_y > 0 else _dash) + '</td>'
            )
        row_style = " style='background:#fffef0'" if i <= 3 else ""
        asl = item["auslastung"]
        asl_color = "#2d7a2d" if asl >= 60 else ("#e67e00" if asl >= 30 else "#cc3333")
        ranking_rows += f'''
                <tr{row_style}>
                    <td class="num" style="font-size:15px;text-align:center;width:52px">{medal}</td>
                    <td>
                        <span style="font-weight:500">{item["name"]}</span>
                        <div class="rank-bar-bg"><div class="rank-bar" style="width:{bar_pct}%"></div></div>
                    </td>
                    <td style="color:#666;font-size:12px">{item["ort"]}</td>
                    {year_cells}
                    <td class="num"><strong style="color:var(--color-primary)">{format_euro(item["total"])}</strong></td>
                    <td class="num" style="color:#555">{int(item["naechte"])}</td>
                    <td class="num" style="color:{asl_color};font-weight:600">{asl:.1f}\u00a0%</td>
                </tr>'''

    yr_range = f"{min(years)}\u2013{max(years)}" if years else ""
    umsatz_ranking_html = f'''
    <div class="chart-container">
        <h3>&#127942; Umsatz-Ranking nach Unterkunft</h3>
        <p style="color:#666;font-size:13px;margin-bottom:20px;">
            Sortiert nach Umsatz (Reisepreis) {current_year} &mdash; h\u00f6chster Umsatz zuerst. Alle {len(ranking_data)} Unterk\u00fcnfte mit Buchungen im Zeitraum {yr_range}.
            Auslastung = gebuchte N\u00e4chte \u00f7 365 Tage je Jahr.
        </p>
        <div style="overflow-x:auto;">
        <table class="prov-table">
            <thead>
                <tr>
                    <th rowspan="2" style="width:52px;text-align:center">Rang</th>
                    <th rowspan="2">Unterkunft</th>
                    <th rowspan="2">Ort</th>
                    {year_header_top}
                    <th rowspan="2" class="num">Gesamt</th>
                    <th rowspan="2" class="num">N\u00e4chte</th>
                    <th rowspan="2" class="num">Auslastung</th>
                </tr>
                <tr>
                    {year_header_sub}
                </tr>
            </thead>
            <tbody>
                {ranking_rows}
            </tbody>
        </table>
        </div>
    </div>'''

    # -----------------------------------------------------------------------
    # Preisliste Tab
    # -----------------------------------------------------------------------
    preisliste_data = data.get("preisliste")
    if preisliste_data and preisliste_data.get("price_lists"):
        pl_year = preisliste_data.get("year", "")
        pl_lists = preisliste_data["price_lists"]

        # Chronologische Reihenfolge der Saisonzeiten (Januar → Dezember)
        SEASON_ORDER = [
            "Winter",
            "Frühling I",
            "Ostern",
            "Frühling II",
            "Frühsommer",
            "Strandzeit I",
            "Strandzeit II",
            "Kranichzeit",
            "Spätherbst",
            "Weihnachten",
            "Silvester - Neujahr",
        ]

        # Collect all unique season names across all price lists
        seen_seasons = set()
        for pl in pl_lists:
            for s in pl.get("season_prices", {}):
                seen_seasons.add(s)

        # Sortierung: bekannte Saisons nach SEASON_ORDER, unbekannte alphabetisch ans Ende
        def _season_sort_key(s):
            try:
                return (0, SEASON_ORDER.index(s))
            except ValueError:
                return (1, s)

        all_seasons = sorted(seen_seasons, key=_season_sort_key)

        # Helper: derive price group label from list name
        def _pg_label(name):
            for stop in [" - ", " Haus", " Barth", " reduz", " spez", " Personen"]:
                idx = name.find(stop)
                if idx > 0:
                    return name[:idx].strip()
            return name.strip()

        # Saisonzeiträume (typische Ostseeliebe-Daten)
        SEASON_DATES = {
            "Winter":              "01.01.–28.02. & 01.11.–19.12.",
            "Frühling I":          "01.03.–11.04.",
            "Frühling II":         "12.04.–16.05.",
            "Frühsommer":          "17.05.–20.06.",
            "Strandzeit I":        "21.06.–01.08.",
            "Strandzeit II":       "02.08.–06.09.",
            "Kranichzeit":         "07.09.–31.10.",
            "Spätherbst":          "01.11.–19.12.",
            "Weihnachten":         "20.12.–26.12.",
            "Silvester - Neujahr": "27.12.–04.01.",
            "Ostern":              "Karwoche–Ostermontag",
        }

        # PG color map – deckt PG-Namen (PG 01 …) UND A-Gruppen (A01 …) und B-Gruppen ab
        _PALETTE = [
            "#FFF2CC","#FCE4D6","#DDEBF7","#E2EFDA","#F4CCCC",
            "#D9EAD3","#CFE2F3","#EAD1DC","#D0E4F5","#F9E4B7",
        ]
        PG_COLORS = {
            "00": "#e8e8e8",
            # Altes PG-Schema
            "PG 01":"#FFF2CC","PG 02":"#FCE4D6","PG 03":"#DDEBF7",
            "PG 04":"#E2EFDA","PG 05":"#F4CCCC","PG 06":"#D9EAD3",
            "PG 07":"#CFE2F3","PG 08":"#FFF2CC","PG 09":"#FCE4D6",
            "PG 10":"#DDEBF7","PG 10.2":"#E2EFDA","PG 11":"#F4CCCC",
            "PG 11.4":"#D9EAD3","PG 12":"#CFE2F3","PG 13":"#FFF2CC",
            "PG 14":"#FCE4D6","PG 15":"#DDEBF7","PG 16":"#E2EFDA",
            "PG 17.4":"#F4CCCC","PG 18":"#D9EAD3","PG 19":"#CFE2F3",
            "PG 20":"#FFF2CC","PG 21":"#FCE4D6",
            # Neues A-Schema (A01–A16)
            "A01":"#FFF2CC","A02":"#FCE4D6","A03":"#DDEBF7",
            "A04":"#E2EFDA","A05":"#F4CCCC","A06":"#D9EAD3",
            "A07":"#CFE2F3","A08":"#EAD1DC","A09":"#D0E4F5",
            "A10":"#F9E4B7","A11":"#FFF2CC","A12":"#FCE4D6",
            "A13":"#DDEBF7","A14":"#E2EFDA","A15":"#F4CCCC","A16":"#D9EAD3",
            # B-Gruppen (Apartmenthaus)
            "B1":"#e8f4e8","B2":"#d4ecd4","B3":"#bfe0bf","B4":"#a8d4a8",
        }

        # Build header – Saisonname + Zeitraum als Tooltip/Zweizeiler
        def _season_th(s):
            dates = SEASON_DATES.get(s, "")
            if dates:
                return (f'<th class="num" title="{dates}" style="cursor:help;">'
                        f'{s}<br><small style="color:#999;font-weight:400">{dates}</small></th>')
            return f'<th class="num">{s}</th>'

        season_headers = "".join(_season_th(s) for s in all_seasons)
        pl_header = f'''
            <tr>
                <th>Preisgruppe</th>
                <th>Unterkunft</th>
                {season_headers}
            </tr>'''

        # Build rows: one row per accommodation
        pl_rows_html = []
        for pl in pl_lists:
            pg = _pg_label(pl["name"])
            bg = PG_COLORS.get(pg, "#f9f9f9")
            sp = pl.get("season_prices", {})
            accoms = pl.get("accommodations", [])
            if not accoms:
                accoms = [{"nr": "", "name": pl["name"], "parking": ""}]

            first = True
            for ac in accoms:
                ac_name = ac.get("name", "")
                price_cells = "".join(
                    '<td class="num">{}</td>'.format(
                        f"{sp[s]:,.0f} \u20ac".replace(",", ".") if s in sp and sp[s] else "\u2013"
                    )
                    for s in all_seasons
                )
                pg_cell = (f'<td rowspan="{len(accoms)}" style="background:{bg};font-weight:600;'
                           f'vertical-align:middle">{pg}</td>') if first else ""
                pl_rows_html.append(
                    f'<tr style="background:{bg if first else ""};">'
                    f'{pg_cell}'
                    f'<td>{ac_name}</td>'
                    f'{price_cells}</tr>'
                )
                first = False

        fetched = preisliste_data.get("fetched_at", "")[:10]
        preisliste_tab_html = f'''
    <div class="chart-container">
        <h3>&#128203; Preislisten {pl_year}</h3>
        <p style="color:#666;font-size:13px;margin-bottom:16px;">
            Saisonpreise pro Nacht je Preisgruppe &mdash; Stand: {fetched} &mdash;
            {len(pl_lists)} Preisgruppen, {sum(len(p.get("accommodations",[]) or [{}]) for p in pl_lists)} Unterk\u00fcnfte
        </p>
        <input type="text" id="plSearch" placeholder="Unterkunft oder Preisgruppe suchen\u2026"
               oninput="filterPL(this.value)"
               style="margin-bottom:14px;padding:8px 12px;border:1px solid #ddd;border-radius:6px;
                      width:300px;font-size:14px;">
        <div style="overflow-x:auto;">
        <table class="prov-table" id="plTable">
            <thead>{pl_header}</thead>
            <tbody>
                {"".join(pl_rows_html)}
            </tbody>
        </table>
        </div>
    </div>
    <script>
    function filterPL(q) {{
        q = q.toLowerCase();
        const rows = document.querySelectorAll('#plTable tbody tr');
        rows.forEach(r => {{
            r.style.display = r.innerText.toLowerCase().includes(q) ? '' : 'none';
        }});
    }}
    </script>'''
    else:
        preisliste_tab_html = '''
    <div class="chart-container">
        <h3>&#128203; Preislisten</h3>
        <p style="color:#888;font-size:14px;">
            Noch keine Preisdaten vorhanden. Die Daten werden beim n\u00e4chsten
            Workflow-Lauf automatisch geladen.
        </p>
    </div>'''

    TABS = [
        ("uebersicht",        "\u00dcbersicht"),
        ("jahresvergleich",   "Jahresvergleich"),
        ("vertriebskanaele",  "Vertriebskan\u00e4le"),
        ("orte",              "Orte"),
        ("zusatzkosten",      "Zusatzkosten"),
        ("provisionen",       "Provisionen"),
        ("umsatz_ranking",    "Umsatz-Ranking"),
        ("preisliste",        "Preisliste"),
        ("apartmenthaeuser",  "Apartmenthaus"),
        ("unterkunft_detail", "Unterkunft Detail"),
    ]
    tab_nav_html = "\n        ".join(
        '<div class="tab{}" data-tab="{}">{}</div>'.format(
            " active" if i == 0 else "", tid, tlabel
        )
        for i, (tid, tlabel) in enumerate(TABS)
    )

    # Tab content dict — keyed by tab id (must match TABS entries above).
    # Each value is a fully-rendered HTML string.
    tab_contents = {
        "uebersicht": (
            bestand_html + "\n" + kpi_html
        ),
        "jahresvergleich": (
            '''<div class="chart-container">
            <h3>\u00dcbernachtungen pro Monat (Jahresvergleich)</h3>
            <div class="chart-wrapper line-chart"><canvas id="monthlyChart"></canvas></div>
        </div>
        <div class="chart-container">
            <h3>Buchungen pro Monat (Vergleichstabelle)</h3>
            ''' + comparison_table + '''
        </div>'''
        ),
        "vertriebskanaele": channel_year_html,
        "orte": _build_orte_tab(orte, years, current_year),
        "zusatzkosten": (
            '''<div class="chart-container">
            <h3>Alle Zusatzkosten \u2013 Detailtabelle</h3>
            ''' + zusatz_detail_table + '''
        </div>'''
        ),
        "provisionen": (
            '''<div class="chart-container">
            <h3>Provisionseinnahmen Ostseeliebe \u2013 nach Unterkunft</h3>
            <p style="color:#666;font-size:13px;margin-bottom:16px;">Miete Vermittler + Zusatzkosten Vermittler = Provision gesamt. Sortiert nach h\u00f6chster Provision.</p>
        </div>
        ''' + prov_tab_html
        ),
        "umsatz_ranking": umsatz_ranking_html,
        "preisliste": preisliste_tab_html,
        "apartmenthaeuser": apartmenthaus_tab_html,
        "unterkunft_detail": (
            '''<div class="chart-container">
            <h3>Unterkunft ausw\u00e4hlen</h3>
            <select id="propSelect" class="prop-select">
                <option value="">-- Bitte Unterkunft w\u00e4hlen --</option>
                ''' + property_options + '''
            </select>
        </div>
        <div id="propDetailContent"></div>'''
        ),
    }

    tab_content_html = "\n\n".join(
        '        <!-- {label} -->\n        <div class="tab-content{active}" id="{tid}">\n{content}\n        </div>'.format(
            label=tlabel,
            active=" active" if i == 0 else "",
            tid=tid,
            content=tab_contents.get(tid, ""),
        )
        for i, (tid, tlabel) in enumerate(TABS)
    )

    html = f'''<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ostseeliebe - Buchungs-Dashboard</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <style>
        /* ================================================================
           Design tokens — change colors / shadows / radii HERE only.
           Do NOT hardcode #0066cc or similar values anywhere else in CSS.
           ================================================================ */
        :root {{
            --color-primary:    #0066cc;
            --color-accent:     #00aaaa;
            --color-danger:     #ff6b6b;
            --color-warning:    #ffa500;
            --color-teal:       #4ecdc4;
            --color-lavender:   #aa96da;
            --color-mint:       #95e1d3;
            --color-green:      #2e7d32;
            --color-bg:         #f5f7fa;
            --color-border:     #e0e0e0;
            --color-text:       #333;
            --color-text-muted: #888;
            --color-table-head-bg: #e8f0fe;
            --shadow-card:      0 2px 8px rgba(0,0,0,0.08);
            --shadow-card-hover: 0 4px 16px rgba(0,0,0,0.12);
            --radius-card:      12px;
            --radius-sm:        8px;
        }}
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, sans-serif;
            background: var(--color-bg);
            color: var(--color-text);
        }}
        .header {{
            background: linear-gradient(135deg, var(--color-primary), var(--color-accent));
            color: white;
            padding: 30px 40px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 28px;
            margin-bottom: 8px;
        }}
        .header .subtitle {{
            font-size: 14px;
            opacity: 0.85;
        }}
        .tabs {{
            display: flex;
            background: white;
            border-bottom: 2px solid #e0e0e0;
            padding: 0 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        .tab {{
            padding: 14px 24px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            color: #666;
            border-bottom: 3px solid transparent;
            transition: all 0.2s;
        }}
        .tab:hover {{
            color: var(--color-primary);
        }}
        .tab.active {{
            color: var(--color-primary);
            border-bottom-color: var(--color-primary);
        }}
        .content {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 30px 20px;
        }}
        .tab-content {{
            display: none;
        }}
        .tab-content.active {{
            display: block;
        }}
        .year-section {{
            margin-bottom: 24px;
        }}
        .year-title {{
            font-size: 20px;
            color: var(--color-primary);
            margin-bottom: 12px;
            padding-bottom: 6px;
            border-bottom: 2px solid var(--color-border);
        }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 16px;
        }}
        .kpi-card {{
            background: white;
            border-radius: 12px;
            padding: 20px;
            text-align: center;
            box-shadow: var(--shadow-card);
            transition: transform 0.2s;
        }}
        .kpi-card:hover {{
            transform: translateY(-2px);
            box-shadow: var(--shadow-card-hover);
        }}
        .kpi-label {{
            font-size: 12px;
            color: var(--color-text-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }}
        .kpi-value {{
            font-size: 20px;
            font-weight: 700;
            color: var(--color-primary);
        }}
        .chart-container {{
            background: white;
            border-radius: 12px;
            padding: 24px;
            box-shadow: var(--shadow-card);
            margin-bottom: 24px;
        }}
        .chart-container h3 {{
            margin-bottom: 16px;
            color: #333;
        }}
        .chart-wrapper {{
            position: relative;
            width: 100%;
        }}
        .chart-wrapper.line-chart {{
            height: 400px;
        }}
        .chart-wrapper.bar-chart {{
            height: 400px;
        }}
        .chart-wrapper.hbar-chart {{
            height: 500px;
        }}
        .chart-wrapper.doughnut-chart {{
            height: 450px;
            max-width: 600px;
            margin: 0 auto;
        }}
        .comparison-table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        .comparison-table th,
        .comparison-table td {{
            padding: 10px 16px;
            text-align: center;
            border-bottom: 1px solid #eee;
            font-size: 14px;
        }}
        .comparison-table th {{
            background: var(--color-primary);
            color: white;
            font-weight: 600;
        }}
        .comparison-table tr:hover {{
            background: #f0f7ff;
        }}
        .total-row {{
            background: var(--color-bg) !important;
        }}
        .total-row td {{
            border-top: 2px solid var(--color-primary);
        }}
        .prov-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            white-space: nowrap;
        }}
        .prov-table th {{
            background: var(--color-primary);
            color: #fff !important;
            font-weight: 600;
            padding: 10px 12px;
            text-align: left;
            position: sticky;
            top: 0;
        }}
        .prov-table th.num {{
            text-align: right;
        }}
        .prov-table td {{
            padding: 7px 12px;
            border-bottom: 1px solid #eee;
        }}
        .prov-table td.num {{
            text-align: right;
            font-variant-numeric: tabular-nums;
        }}
        .prov-table tr:hover {{
            background: #f0f7ff;
        }}
        .prov-total {{
            background: #f0f4f8 !important;
            border-top: 2px solid var(--color-primary);
        }}
        .rank-bar-bg {{
            background: #eef2f7;
            height: 4px;
            border-radius: 2px;
            margin-top: 4px;
            width: 160px;
        }}
        .rank-bar {{
            height: 4px;
            border-radius: 2px;
            background: var(--color-primary);
        }}
        .pk-value.green {{
            color: #2e7d32 !important;
        }}
        .prop-select {{
            width: 100%;
            padding: 12px 16px;
            font-size: 16px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            background: white;
            cursor: pointer;
            outline: none;
            transition: border-color 0.2s;
        }}
        .prop-select:focus {{
            border-color: var(--color-primary);
        }}
        .prop-detail-grid {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 12px;
            margin-bottom: 16px;
        }}
        .prop-kpi {{
            background: white;
            border-radius: 10px;
            padding: 16px;
            text-align: center;
            box-shadow: 0 1px 6px rgba(0,0,0,0.07);
        }}
        .prop-kpi .pk-label {{
            font-size: 11px;
            color: var(--color-text-muted);
            text-transform: uppercase;
            letter-spacing: 0.4px;
            margin-bottom: 6px;
        }}
        .prop-kpi .pk-value {{
            font-size: 18px;
            font-weight: 700;
            color: var(--color-primary);
        }}
        .prop-kpi .pk-value.green {{
            color: var(--color-green);
        }}
        .prop-section {{
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: var(--shadow-card);
            margin-bottom: 20px;
        }}
        .prop-section h4 {{
            margin-bottom: 12px;
            color: #333;
            font-size: 16px;
        }}
        .prop-channel-table,
        .prop-zk-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        .prop-channel-table th,
        .prop-zk-table th {{
            background: var(--color-table-head-bg);
            color: var(--color-primary);
            font-weight: 600;
            padding: 8px 10px;
            text-align: left;
        }}
        .prop-channel-table td,
        .prop-zk-table td {{
            padding: 6px 10px;
            border-bottom: 1px solid #f0f0f0;
        }}
        .prop-channel-table .num,
        .prop-zk-table .num {{
            text-align: right;
            font-variant-numeric: tabular-nums;
        }}
        .prop-zk-table .sub {{
            color: #888;
            font-size: 12px;
        }}
        .zusatz-summary {{
            margin-top: 16px;
        }}
        .zusatz-table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 1px 4px rgba(0,0,0,0.06);
            font-size: 13px;
        }}
        .zusatz-table th {{
            background: var(--color-table-head-bg);
            color: var(--color-primary);
            font-weight: 600;
            padding: 8px 12px;
            text-align: left;
        }}
        .zusatz-table td {{
            padding: 6px 12px;
            border-bottom: 1px solid #f0f0f0;
        }}
        .zusatz-table .zk-num {{
            text-align: right;
            font-variant-numeric: tabular-nums;
        }}
        .zusatz-table .zk-sub {{
            color: #888;
            font-size: 12px;
        }}
        .zusatz-table th.zk-num {{
            text-align: right;
        }}
        .zusatz-table th.zk-sub {{
            color: #5588bb;
        }}
        .zusatz-table .zk-total {{
            background: #f5f7fa;
            border-top: 2px solid var(--color-primary);
        }}
        .zusatz-detail-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
            white-space: nowrap;
        }}
        .zusatz-detail-table th {{
            background: var(--color-primary);
            color: white;
            font-weight: 600;
            padding: 8px 10px;
            text-align: left;
            position: sticky;
            top: 0;
        }}
        .zusatz-detail-table td {{
            padding: 6px 10px;
            border-bottom: 1px solid #eee;
        }}
        .zusatz-detail-table .zk-num {{
            text-align: right;
            font-variant-numeric: tabular-nums;
        }}
        .zusatz-detail-table .zk-sub {{
            color: #888;
        }}
        .zusatz-detail-table th.zk-num {{
            text-align: right;
        }}
        .zusatz-detail-table th.zk-sub {{
            color: #aaccff;
        }}
        .zusatz-detail-table .zk-total {{
            background: #f0f4f8;
            border-top: 2px solid var(--color-primary);
        }}
        .zusatz-detail-table tr:hover {{
            background: #f0f7ff;
        }}
        @media (max-width: 900px) {{
            .kpi-grid {{
                grid-template-columns: repeat(2, 1fr);
            }}
            .tabs {{
                flex-wrap: wrap;
            }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Ostseeliebe - Buchungs-Dashboard</h1>
        <div class="subtitle">Aktualisiert am {update_date}</div>
    </div>

    <div class="tabs">
        {tab_nav_html}
    </div>

    <div class="content">
        {tab_content_html}
    </div>

    <script>
        // Tab switching
        document.querySelectorAll('.tab').forEach(function(tab) {{
            tab.addEventListener('click', function() {{
                document.querySelectorAll('.tab').forEach(function(t) {{ t.classList.remove('active'); }});
                document.querySelectorAll('.tab-content').forEach(function(c) {{ c.classList.remove('active'); }});
                tab.classList.add('active');
                document.getElementById(tab.dataset.tab).classList.add('active');
            }});
        }});

        // Chart.js defaults
        Chart.defaults.font.family = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif";

        // Monthly Line Chart
        new Chart(document.getElementById('monthlyChart'), {{
            type: 'line',
            data: {{
                labels: {json_month_names},
                datasets: {json_line_datasets}
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ position: 'top' }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{ display: true, text: 'N\u00e4chte' }}
                    }}
                }}
            }}
        }});

        // Sales Channels Doughnut Charts (per year)
        {channel_chart_js}

        // Locations Bar Chart
        new Chart(document.getElementById('ortChart'), {{
            type: 'bar',
            data: {{
                labels: {ort_labels},
                datasets: [{{
                    label: 'Buchungen',
                    data: {ort_values},
                    backgroundColor: {ort_colors},
                    borderRadius: 6
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ display: false }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{ display: true, text: 'Anzahl Buchungen' }}
                    }}
                }}
            }}
        }});

        // --- Unterkunft Detail ---
        var propData = {json_property_data};
        var allYears = {json_years};
        var propChartInstance = null;

        function fmtEuro(n) {{
            return new Intl.NumberFormat('de-DE', {{style:'currency',currency:'EUR'}}).format(n);
        }}
        function fmtNum(n, d) {{
            return new Intl.NumberFormat('de-DE', {{minimumFractionDigits:d||0, maximumFractionDigits:d||0}}).format(n);
        }}

        function renderProperty(name) {{
            var container = document.getElementById('propDetailContent');
            if (!name || !propData[name]) {{
                container.innerHTML = '<p style="color:#888;padding:20px;">Bitte eine Unterkunft ausw\u00e4hlen.</p>';
                return;
            }}
            var p = propData[name];
            var html = '<h3 style="margin:20px 0 6px;font-size:22px;color:#0066cc;">' + name + '</h3>';
            html += '<p style="color:#666;margin-bottom:20px;">Ort: ' + p.ort + '</p>';

            allYears.forEach(function(y) {{
                var yd = p.years[y];
                if (!yd || yd.buchungen === 0) return;
                html += '<div class="prop-section">';
                html += '<h4 style="color:#0066cc;border-bottom:2px solid #e0e0e0;padding-bottom:6px;">' + y + '</h4>';

                // KPI Grid
                html += '<div class="prop-detail-grid">';
                html += '<div class="prop-kpi"><div class="pk-label">Buchungen</div><div class="pk-value">' + yd.buchungen + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">N\u00e4chte</div><div class="pk-value">' + fmtNum(yd.naechte) + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">Reisepreis</div><div class="pk-value">' + fmtEuro(yd.reisepreis) + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">Miete gesamt</div><div class="pk-value">' + fmtEuro(yd.miete_gesamt) + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">Miete Eigent\u00fcmer</div><div class="pk-value green">' + fmtEuro(yd.miete_eigentuemer) + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">Provision (Verm.)</div><div class="pk-value" style="color:#e65100;">' + fmtEuro(yd.miete_vermittler) + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">Prov.satz</div><div class="pk-value" style="color:#e65100;">' + fmtNum(yd.provision_pct, 1) + ' %</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">\u00d8 Preis/Nacht</div><div class="pk-value">' + fmtEuro(yd.avg_preis_nacht) + '</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">\u00d8 Aufenthalt</div><div class="pk-value">' + fmtNum(yd.avg_aufenthalt, 1) + ' N.</div></div>';
                html += '<div class="prop-kpi"><div class="pk-label">Auslastung</div><div class="pk-value">' + fmtNum(yd.belegung_pct, 1) + ' %</div></div>';
                html += '</div>';

                // Vertriebskanäle
                if (yd.channels && yd.channels.length > 0) {{
                    html += '<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:12px;">';
                    html += '<div>';
                    html += '<h4 style="font-size:14px;margin-bottom:8px;">Vertriebskan\u00e4le</h4>';
                    html += '<table class="prop-channel-table"><thead><tr><th>Kanal</th><th class="num">Buchungen</th></tr></thead><tbody>';
                    yd.channels.forEach(function(ch) {{
                        html += '<tr><td>' + (ch[0] || '(leer)') + '</td><td class="num">' + ch[1] + '</td></tr>';
                    }});
                    html += '</tbody></table></div>';

                    // Zusatzkosten
                    html += '<div>';
                    html += '<h4 style="font-size:14px;margin-bottom:8px;">Zusatzkosten</h4>';
                    if (yd.zusatzkosten && yd.zusatzkosten.length > 0) {{
                        html += '<table class="prop-zk-table"><thead><tr><th>Kategorie</th><th class="num">Gesamt</th><th class="num sub">Vermittler</th><th class="num sub">Eigent\u00fcmer</th></tr></thead><tbody>';
                        var zkTotal = 0, zkV = 0, zkE = 0;
                        yd.zusatzkosten.forEach(function(z) {{
                            html += '<tr><td>' + z.name + '</td><td class="num">' + fmtEuro(z.gesamt) + '</td><td class="num sub">' + fmtEuro(z.vermittler) + '</td><td class="num sub">' + fmtEuro(z.eigentuemer) + '</td></tr>';
                            zkTotal += z.gesamt; zkV += z.vermittler; zkE += z.eigentuemer;
                        }});
                        html += '<tr style="background:#f5f7fa;border-top:2px solid #0066cc;"><td><strong>Summe</strong></td><td class="num"><strong>' + fmtEuro(zkTotal) + '</strong></td><td class="num sub"><strong>' + fmtEuro(zkV) + '</strong></td><td class="num sub"><strong>' + fmtEuro(zkE) + '</strong></td></tr>';
                        html += '</tbody></table>';
                    }} else {{
                        html += '<p style="color:#aaa;">Keine Zusatzkosten</p>';
                    }}
                    html += '</div></div>';
                }}

                html += '</div>'; // prop-section
            }});

            // Jahresvergleich mini chart
            html += '<div class="prop-section"><h4>Jahresvergleich \u2013 Reisepreis</h4><canvas id="propYearChart" style="max-height:300px;"></canvas></div>';
            container.innerHTML = html;

            // Render mini chart
            if (propChartInstance) propChartInstance.destroy();
            var ctx = document.getElementById('propYearChart');
            if (ctx) {{
                var labels = []; var rpData = []; var meData = [];
                allYears.forEach(function(y) {{
                    var yd = p.years[y];
                    if (yd && yd.buchungen > 0) {{
                        labels.push(y);
                        rpData.push(yd.reisepreis);
                        meData.push(yd.miete_eigentuemer);
                    }}
                }});
                propChartInstance = new Chart(ctx, {{
                    type: 'bar',
                    data: {{
                        labels: labels,
                        datasets: [
                            {{ label: 'Reisepreis', data: rpData, backgroundColor: '#0066cc', borderRadius: 4 }},
                            {{ label: 'Miete Eigent\u00fcmer', data: meData, backgroundColor: '#00aaaa', borderRadius: 4 }}
                        ]
                    }},
                    options: {{
                        responsive: true,
                        plugins: {{
                            legend: {{ position: 'top' }},
                            tooltip: {{
                                callbacks: {{
                                    label: function(ctx) {{ return ctx.dataset.label + ': ' + fmtEuro(ctx.raw); }}
                                }}
                            }}
                        }},
                        scales: {{
                            y: {{
                                beginAtZero: true,
                                ticks: {{ callback: function(v) {{ return fmtEuro(v); }} }}
                            }}
                        }}
                    }}
                }});
            }}
        }}

        document.getElementById('propSelect').addEventListener('change', function() {{
            renderProperty(this.value);
        }});
    </script>
</body>
</html>'''
    return html


def generate_property_html(prop_name, pdata, years):
    """Generate a standalone HTML file for a single property."""
    update_date = datetime.now().strftime("%d.%m.%Y %H:%M")

    year_sections = ""
    chart_labels = []
    chart_rp = []
    chart_me = []

    for y in years:
        yd = pdata["years"].get(y)
        if not yd or yd["buchungen"] == 0:
            continue

        chart_labels.append(y)
        chart_rp.append(round(yd["reisepreis"], 2))
        chart_me.append(round(yd["miete_eigentuemer"], 2))
        chart_mv = round(yd.get("miete_vermittler", 0), 2)

        # Channels table
        ch_rows = ""
        for ch in yd["channels"]:
            ch_name = ch[0] if isinstance(ch, (list, tuple)) else ch
            ch_count = ch[1] if isinstance(ch, (list, tuple)) else 0
            ch_rows += f'<tr><td>{ch_name or "(leer)"}</td><td class="num">{ch_count}</td></tr>'

        # Zusatzkosten table
        zk_rows = ""
        zk_total_g = zk_total_v = zk_total_e = 0
        for z in yd["zusatzkosten"]:
            zk_rows += f'<tr><td>{z["name"]}</td><td class="num">{format_euro(z["gesamt"])}</td><td class="num sub">{format_euro(z["vermittler"])}</td><td class="num sub">{format_euro(z["eigentuemer"])}</td></tr>'
            zk_total_g += z["gesamt"]
            zk_total_v += z["vermittler"]
            zk_total_e += z["eigentuemer"]
        if zk_rows:
            zk_rows += f'<tr style="background:#f5f7fa;border-top:2px solid #0066cc;"><td><strong>Summe</strong></td><td class="num"><strong>{format_euro(zk_total_g)}</strong></td><td class="num sub"><strong>{format_euro(zk_total_v)}</strong></td><td class="num sub"><strong>{format_euro(zk_total_e)}</strong></td></tr>'

        year_sections += f'''
        <div class="section">
            <h2>{y}</h2>
            <div class="kpi-grid">
                <div class="kpi"><div class="kl">Buchungen</div><div class="kv">{yd["buchungen"]}</div></div>
                <div class="kpi"><div class="kl">N\u00e4chte</div><div class="kv">{format_german_number(yd["naechte"], 0)}</div></div>
                <div class="kpi"><div class="kl">Reisepreis</div><div class="kv">{format_euro(yd["reisepreis"])}</div></div>
                <div class="kpi"><div class="kl">Miete gesamt</div><div class="kv">{format_euro(yd["miete_gesamt"])}</div></div>
                <div class="kpi"><div class="kl">Miete Eigent\u00fcmer</div><div class="kv green">{format_euro(yd["miete_eigentuemer"])}</div></div>
                <div class="kpi"><div class="kl">Provision (Verm.)</div><div class="kv" style="color:#e65100;">{format_euro(yd.get("miete_vermittler", 0))}</div></div>
                <div class="kpi"><div class="kl">Prov.satz</div><div class="kv" style="color:#e65100;">{format_german_number(yd.get("provision_pct", 0), 1)} %</div></div>
                <div class="kpi"><div class="kl">\u00d8 Preis/Nacht</div><div class="kv">{format_euro(yd["avg_preis_nacht"])}</div></div>
                <div class="kpi"><div class="kl">\u00d8 Aufenthalt</div><div class="kv">{format_german_number(yd["avg_aufenthalt"], 1)} N.</div></div>
                <div class="kpi"><div class="kl">Auslastung</div><div class="kv">{format_german_number(yd["belegung_pct"], 1)} %</div></div>
            </div>
            <div class="two-col">
                <div>
                    <h3>Vertriebskan\u00e4le</h3>
                    <table><thead><tr><th>Kanal</th><th class="num">Buchungen</th></tr></thead><tbody>{ch_rows}</tbody></table>
                </div>
                <div>
                    <h3>Zusatzkosten</h3>
                    {"<table><thead><tr><th>Kategorie</th><th class='num'>Gesamt</th><th class='num sub'>Vermittler</th><th class='num sub'>Eigent.</th></tr></thead><tbody>" + zk_rows + "</tbody></table>" if zk_rows else "<p style='color:#aaa;'>Keine Zusatzkosten</p>"}
                </div>
            </div>
        </div>'''

    json_labels = json.dumps(chart_labels)
    json_rp = json.dumps(chart_rp)
    json_me = json.dumps(chart_me)

    return f'''<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{prop_name} \u2013 Umsatz\u00fcbersicht</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <style>
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#f5f7fa; color:#333; }}
        .header {{ background:linear-gradient(135deg,#0066cc,#00aaaa); color:white; padding:30px 40px; }}
        .header h1 {{ font-size:24px; margin-bottom:4px; }}
        .header .sub {{ opacity:0.85; font-size:13px; }}
        .container {{ max-width:1100px; margin:0 auto; padding:24px 20px; }}
        .section {{ background:white; border-radius:12px; padding:24px; box-shadow:0 2px 8px rgba(0,0,0,0.08); margin-bottom:20px; }}
        .section h2 {{ color:#0066cc; font-size:20px; border-bottom:2px solid #e0e0e0; padding-bottom:6px; margin-bottom:16px; }}
        .kpi-grid {{ display:grid; grid-template-columns:repeat(5,1fr); gap:12px; margin-bottom:16px; }}
        .kpi {{ background:#f8fafc; border-radius:10px; padding:14px; text-align:center; }}
        .kl {{ font-size:11px; color:#888; text-transform:uppercase; letter-spacing:.4px; margin-bottom:4px; }}
        .kv {{ font-size:17px; font-weight:700; color:#0066cc; }}
        .kv.green {{ color:#2e7d32; }}
        .two-col {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
        h3 {{ font-size:14px; margin-bottom:8px; color:#555; }}
        table {{ width:100%; border-collapse:collapse; font-size:13px; }}
        th {{ background:#e8f0fe; color:#0066cc; font-weight:600; padding:8px 10px; text-align:left; }}
        td {{ padding:6px 10px; border-bottom:1px solid #f0f0f0; }}
        .num {{ text-align:right; font-variant-numeric:tabular-nums; }}
        .sub {{ color:#888; font-size:12px; }}
        .chart-section {{ background:white; border-radius:12px; padding:24px; box-shadow:0 2px 8px rgba(0,0,0,0.08); margin-bottom:20px; }}
        .chart-section h3 {{ font-size:16px; color:#333; margin-bottom:12px; }}
        @media(max-width:800px) {{ .kpi-grid {{ grid-template-columns:repeat(2,1fr); }} .two-col {{ grid-template-columns:1fr; }} }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{prop_name}</h1>
        <div class="sub">{pdata["ort"]} \u2014 Umsatz\u00fcbersicht \u2014 Stand {update_date}</div>
    </div>
    <div class="container">
        <div class="chart-section">
            <h3>Jahresvergleich \u2013 Reisepreis & Miete Eigent\u00fcmer</h3>
            <canvas id="yearChart" style="max-height:300px;"></canvas>
        </div>
        {year_sections}
    </div>
    <script>
        new Chart(document.getElementById('yearChart'), {{
            type: 'bar',
            data: {{
                labels: {json_labels},
                datasets: [
                    {{ label: 'Reisepreis', data: {json_rp}, backgroundColor: '#0066cc', borderRadius: 4 }},
                    {{ label: 'Miete Eigent\u00fcmer', data: {json_me}, backgroundColor: '#00aaaa', borderRadius: 4 }}
                ]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{ position: 'top' }},
                    tooltip: {{ callbacks: {{ label: function(c) {{ return c.dataset.label + ': ' + new Intl.NumberFormat('de-DE',{{style:'currency',currency:'EUR'}}).format(c.raw); }} }} }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        ticks: {{ callback: function(v) {{ return new Intl.NumberFormat('de-DE',{{style:'currency',currency:'EUR',maximumFractionDigits:0}}).format(v); }} }}
                    }}
                }}
            }}
        }});
    </script>
</body>
</html>'''


def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CSV
    out_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT

    if not os.path.exists(csv_path):
        print(f"Fehler: CSV-Datei nicht gefunden: {csv_path}")
        sys.exit(1)

    # Objektstammdaten (optional)
    stamm_path = os.path.join(os.path.dirname(csv_path), "Objektstammdaten.xlsx")
    if len(sys.argv) > 3:
        stamm_path = sys.argv[3]

    # Preislisten-JSON (optional, neben der CSV)
    preisliste_path = os.path.join(os.path.dirname(os.path.abspath(csv_path)), "preisliste_data.json")
    preisliste_data = None
    if os.path.exists(preisliste_path):
        try:
            with open(preisliste_path, encoding="utf-8") as _pf:
                preisliste_data = json.load(_pf)
            print(f"  Preislisten geladen: {len(preisliste_data.get('price_lists', []))} Preisgruppen")
        except Exception as _pe:
            print(f"  ⚠️  Preislisten-JSON fehlerhaft: {_pe}")
    else:
        print(f"  ℹ️  Keine Preislisten-Datei ({preisliste_path})")

    print(f"Lese Buchungen aus: {csv_path}")
    bookings = read_bookings(csv_path)
    print(f"  {len(bookings)} Buchungen gelesen (Status=Buchung)")

    stammdaten = read_objektstammdaten(stamm_path)
    if stammdaten:
        t = stammdaten["totals"]
        print(f"  Objektstammdaten: {t['count']} Unterkünfte, {t['wohnflaeche']} m², "
              f"{t['schlafzimmer']} Schlafzimmer, {t['max_personen']} Schlafplätze")
    else:
        print("  Objektstammdaten nicht gefunden oder openpyxl nicht installiert – überspringe Bestandskennzahlen")

    data = compute_data(bookings)
    data["stammdaten"] = stammdaten
    data["preisliste"] = preisliste_data
    print(f"  Jahre: {data['years']}")
    for y in data["years"]:
        k = data["kpis"][y]
        print(f"    {y}: {k['buchungen']} Buchungen, {k['naechte']:.0f} Naechte, "
              f"Reisepreis {k['reisepreis']:.2f}, Miete ges. {k['miete_gesamt']:.2f}, "
              f"Miete Eig. {k['miete_eigentuemer']:.2f}")

    html = generate_html(data)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\nDashboard geschrieben: {out_path}")
    print(f"  Dateigroesse: {os.path.getsize(out_path):,} Bytes")

    # Generate individual property HTML files
    out_dir = os.path.dirname(out_path)
    prop_dir = os.path.join(out_dir, "unterkuenfte")
    os.makedirs(prop_dir, exist_ok=True)

    property_data = data["property_data"]
    years = data["years"]
    count = 0
    for prop_name, pdata in sorted(property_data.items()):
        # Skip properties with no bookings
        if all(pdata["years"].get(y, {}).get("buchungen", 0) == 0 for y in years):
            continue
        # Safe filename
        safe_name = prop_name.replace("/", "-").replace("\\", "-").replace(" ", "_")
        safe_name = "".join(c for c in safe_name if c.isalnum() or c in "-_").strip("_")
        if not safe_name:
            safe_name = f"unterkunft_{count}"
        filepath = os.path.join(prop_dir, f"{safe_name}.html")
        prop_html = generate_property_html(prop_name, pdata, years)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(prop_html)
        count += 1

    print(f"\n{count} Unterkunft-Einzelseiten geschrieben: {prop_dir}/")


if __name__ == "__main__":
    main()
