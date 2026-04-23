#!/usr/bin/env python3
"""
fetch_property_data.py
======================
Loggt sich in das Contao-Backend ein, liest alle Unterkunft-Stammdaten aus
dem fewo_objekte-Modul und schreibt Objektstammdaten.xlsx.

Die erzeugte Excel-Datei ist kompatibel mit read_objektstammdaten() in gen_dash.py:
    Zeile 1–3 : Kopfzeilen (werden uebersprungen, min_row=4)
    Zeile 4+  : Datenzeilen
    Spalte A (0): Name
    Spalte B (1): Objekt-Nr
    Spalte C (2): (reserviert)
    Spalte D (3): Ort
    Spalte E (4): Wohnflaeche (m2)
    Spalte F (5): Zimmer
    Spalte G (6): Schlafzimmer
    Spalte H (7): Badzimmer
    Spalte I (8): Schlafplaetze (max. Personen)
    Spalte J (9): (reserviert)
    Spalte K (10): (reserviert)
    Spalte L (11): Sauna  ("Ja" / "-")
    Spalte M (12): Hund   ("Ja" / "-")
    Spalte N (13): Kamin  ("Ja" / "-")

Benoetigt folgende Umgebungsvariablen (als GitHub Secrets hinterlegt):
    CONTAO_URL      z.B. https://www.ostseeliebe-ferienwohnungen.de
    CONTAO_USER     Benutzername fuer Contao-Login
    CONTAO_PASS     Passwort fuer Contao-Login
"""

import os
import re
import sys
import argparse
from pathlib import Path

try:
    import requests
except ImportError:
    print("requests fehlt. Bitte: pip install requests")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("beautifulsoup4 fehlt. Bitte: pip install beautifulsoup4")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl fehlt. Bitte: pip install openpyxl")
    sys.exit(1)


BASE_URL = os.environ.get("CONTAO_URL", "https://www.ostseeliebe-ferienwohnungen.de")
USERNAME = os.environ.get("CONTAO_USER", "")
PASSWORD = os.environ.get("CONTAO_PASS", "")
OUT_XLSX = "Objektstammdaten.xlsx"

LABEL_MAP = {
    "name":         ["name", "bezeichnung", "objektname", "title", "titel"],
    "objnr":        ["objekt-nr", "objektnr", "objnr", "nr", "number", "alias"],
    "ort":          ["ort", "city", "location", "standort", "gemeinde"],
    "wohnflaeche":  ["wohnflaeche", "flaeche", "qm", "wohnfl", "living"],
    "zimmer":       ["zimmer", "raeume", "rooms"],
    "schlafzimmer": ["schlafzimmer", "schlafraeume", "bedrooms", "schlafraum"],
    "badzimmer":    ["badzimmer", "badezimmer", "baeder", "bathrooms", "bad"],
    "max_personen": ["schlafplaetze", "personen", "persons", "belegung",
                     "max. personen", "maxpersonen", "kapazitaet", "betten"],
    "sauna":        ["sauna"],
    "hund":         ["hund", "haustier", "hunde", "pet", "pets", "haustiere"],
    "kamin":        ["kamin", "fireplace", "offener kamin", "kaminofen"],
}


def _match_label(label_text, field_key):
    lt = label_text.lower().strip().rstrip(":")
    for alias in LABEL_MAP[field_key]:
        if alias in lt:
            return True
    return False


def get_request_token(session, url):
    resp = session.get(url, timeout=30)
    resp.raise_for_status()
    m = re.search(r'name="REQUEST_TOKEN"\s+value="([^"]+)"', resp.text)
    if not m:
        m = re.search(r'<meta name="request-token" content="([^"]+)"', resp.text)
    if not m:
        raise RuntimeError(f"REQUEST_TOKEN nicht gefunden auf: {url}")
    return m.group(1)


def contao_login(session):
    login_url = f"{BASE_URL}/contao/login"
    print(f"Login bei {BASE_URL} ...")
    token = get_request_token(session, login_url)
    resp = session.post(login_url, data={
        "REQUEST_TOKEN": token,
        "username":      USERNAME,
        "password":      PASSWORD,
    }, allow_redirects=True, timeout=30)
    if 'name="username"' in resp.text:
        print("Login fehlgeschlagen")
        return False
    print("Login erfolgreich")
    return True


def get_property_edit_links(session):
    list_url = f"{BASE_URL}/contao?do=fewo_objekte"
    print(f"Lade Objektliste von {list_url} ...")
    resp = session.get(list_url, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    edit_links = []
    seen_ids = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if ("act=edit" in href or "act=show" in href) and "id=" in href:
            id_match = re.search(r"[?&]id=(\d+)", href)
            if id_match:
                prop_id = id_match.group(1)
                if prop_id not in seen_ids:
                    seen_ids.add(prop_id)
                    full_url = href if href.startswith("http") else f"{BASE_URL}/{href.lstrip('/')}"
                    edit_links.append((prop_id, full_url))
    if not edit_links:
        for tr in soup.find_all("tr", id=re.compile(r"row_\d+")):
            id_match = re.search(r"row_(\d+)", tr.get("id", ""))
            if id_match:
                prop_id = id_match.group(1)
                if prop_id not in seen_ids:
                    seen_ids.add(prop_id)
                    rt_link = soup.find("a", href=re.compile(r"rt="))
                    rt = ""
                    if rt_link:
                        rt_match = re.search(r"rt=([^&]+)", rt_link["href"])
                        rt = rt_match.group(1) if rt_match else ""
                    url = f"{BASE_URL}/contao?do=fewo_objekte&act=edit&id={prop_id}"
                    if rt:
                        url += f"&rt={rt}"
                    edit_links.append((prop_id, url))
    print(f"   {len(edit_links)} Objekte gefunden")
    return edit_links


def parse_property_from_edit_page(session, prop_id, edit_url):
    resp = session.get(edit_url, timeout=30)
    if resp.status_code != 200:
        return {}
    soup = BeautifulSoup(resp.text, "html.parser")
    prop = {
        "id": prop_id, "name": "", "objnr": "", "ort": "",
        "wohnflaeche": 0, "zimmer": 0, "schlafzimmer": 0,
        "badzimmer": 0, "max_personen": 0,
        "sauna": False, "hund": False, "kamin": False,
    }
    for widget in soup.find_all(class_=re.compile(r"\bwidget\b")):
        label_el = widget.find("label")
        if not label_el:
            continue
        label_text = label_el.get_text(strip=True)
        checkbox = widget.find("input", {"type": "checkbox"})
        if checkbox:
            checked = checkbox.get("checked") is not None or checkbox.get("value") == "1"
            for key in ("sauna", "hund", "kamin"):
                if _match_label(label_text, key):
                    prop[key] = checked
            continue
        inp = widget.find("input", {"type": ["text", "number", None]})
        if not inp:
            inp = widget.find("select")
            val = inp.find("option", selected=True).get_text(strip=True) if (inp and inp.find("option", selected=True)) else ""
        else:
            val = inp.get("value", "").strip()
        for key in ("name", "objnr", "ort", "wohnflaeche", "zimmer", "schlafzimmer", "badzimmer", "max_personen"):
            if _match_label(label_text, key):
                if key in ("wohnflaeche", "zimmer", "schlafzimmer", "badzimmer", "max_personen"):
                    try:
                        prop[key] = float(val.replace(",", ".")) if val else 0
                    except ValueError:
                        prop[key] = 0
                else:
                    if not prop[key]:
                        prop[key] = val
    if not prop["name"]:
        for row in soup.find_all("tr"):
            cells = row.find_all(["th", "td"])
            if len(cells) >= 2:
                label_text = cells[0].get_text(strip=True)
                val = cells[1].get_text(strip=True)
                for key in ("name", "objnr", "ort", "wohnflaeche", "zimmer", "schlafzimmer", "badzimmer", "max_personen"):
                    if _match_label(label_text, key):
                        if key in ("wohnflaeche", "zimmer", "schlafzimmer", "badzimmer", "max_personen"):
                            try:
                                prop[key] = float(val.replace(",", ".")) if val else 0
                            except ValueError:
                                prop[key] = 0
                        elif not prop[key]:
                            prop[key] = val
    if not prop["objnr"]:
        prop["objnr"] = prop_id
    return prop


def parse_property_list_page(session):
    list_url = f"{BASE_URL}/contao?do=fewo_objekte"
    resp = session.get(list_url, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    header_row = soup.find("tr", class_=re.compile("header|thead")) or                  (soup.find("thead") and soup.find("thead").find("tr"))
    col_map = {}
    if header_row:
        for i, th in enumerate(header_row.find_all(["th", "td"])):
            label = th.get_text(strip=True)
            for key in LABEL_MAP:
                if _match_label(label, key):
                    col_map[key] = i
    properties = []
    tbody = soup.find("tbody") or soup
    for tr in tbody.find_all("tr"):
        if tr.get("class") and any("header" in c for c in tr.get("class", [])):
            continue
        cells = tr.find_all("td")
        if not cells:
            continue
        prop = {k: "" for k in ("id", "name", "objnr", "ort")}
        prop.update({k: 0 for k in ("wohnflaeche", "zimmer", "schlafzimmer", "badzimmer", "max_personen")})
        prop.update({k: False for k in ("sauna", "hund", "kamin")})
        for a in tr.find_all("a", href=True):
            id_match = re.search(r"[?&]id=(\d+)", a["href"])
            if id_match:
                prop["id"] = id_match.group(1)
                break
        if col_map:
            for key, idx in col_map.items():
                if idx < len(cells):
                    val = cells[idx].get_text(strip=True)
                    if key in ("sauna", "hund", "kamin"):
                        prop[key] = val.lower() in ("ja", "yes", "1", "true")
                    elif key in ("wohnflaeche", "zimmer", "schlafzimmer", "badzimmer", "max_personen"):
                        try:
                            prop[key] = float(val.replace(",", ".")) if val else 0
                        except ValueError:
                            prop[key] = 0
                    else:
                        prop[key] = val
        else:
            if len(cells) > 0:
                prop["name"] = cells[0].get_text(strip=True)
            if len(cells) > 1:
                prop["objnr"] = cells[1].get_text(strip=True)
        if prop["id"] or prop["name"]:
            properties.append(prop)
    return properties


def write_xlsx(properties, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Objektstammdaten"
    col_widths = [35, 12, 5, 18, 12, 10, 14, 12, 14, 5, 5, 10, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = "Objektstammdaten - Ostseeliebe Ferienwohnungen"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[2].height = 6
    headers = ["Name", "Objekt-Nr", "-", "Ort", "Wohnflaeche (m2)", "Zimmer",
               "Schlafzimmer", "Badzimmer", "Schlafplaetze", "-", "-", "Sauna", "Hund", "Kamin"]
    header_fill = PatternFill("solid", fgColor="2E75B6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=hdr)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 30
    ja_fill = PatternFill("solid", fgColor="E2EFDA")
    for r, prop in enumerate(properties, 4):
        row_data = [
            prop.get("name", ""), prop.get("objnr", ""), "",
            prop.get("ort", ""), prop.get("wohnflaeche", 0) or 0,
            prop.get("zimmer", 0) or 0, prop.get("schlafzimmer", 0) or 0,
            prop.get("badzimmer", 0) or 0, prop.get("max_personen", 0) or 0,
            "", "", "Ja" if prop.get("sauna") else "-",
            "Ja" if prop.get("hund") else "-", "Ja" if prop.get("kamin") else "-",
        ]
        bg = "FFFFFF" if r % 2 == 0 else "F5F5F5"
        row_fill = PatternFill("solid", fgColor=bg)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if col in (5, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, float) and val == int(val):
                    cell.value = int(val)
            if col in (12, 13, 14):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = ja_fill if val == "Ja" else row_fill
            else:
                cell.fill = row_fill
        ws.row_dimensions[r].height = 18
    ws.auto_filter.ref = f"A3:N{3 + len(properties)}"
    ws.freeze_panes = "A4"
    wb.save(out_path)
    print(f"Gespeichert: {out_path} ({len(properties)} Objekte)")


def main():
    parser = argparse.ArgumentParser(description="Liest Objektstammdaten aus Contao")
    parser.add_argument("--out", default=OUT_XLSX)
    parser.add_argument("--full", action="store_true")
    args = parser.parse_args()
    if not USERNAME or not PASSWORD:
        print("CONTAO_USER und CONTAO_PASS muessen als Umgebungsvariablen gesetzt sein.")
        sys.exit(1)
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; OstseeliebeBot/1.0)"})
    if not contao_login(session):
        sys.exit(1)
    print("Schritt 1: Listenansicht parsen ...")
    properties = parse_property_list_page(session)
    edit_links = get_property_edit_links(session)
    list_has_details = any(
        p.get("wohnflaeche") or p.get("schlafzimmer") or p.get("sauna")
        for p in properties
    )
    if args.full or not list_has_details or not properties:
        print(f"Schritt 2: {len(edit_links)} Edit-Seiten abrufen ...")
        properties = []
        for i, (prop_id, url) in enumerate(edit_links, 1):
            print(f"  [{i}/{len(edit_links)}] ID {prop_id} ...", end=" ", flush=True)
            prop = parse_property_from_edit_page(session, prop_id, url)
            if prop:
                properties.append(prop)
                print(f"OK {prop.get('name', '?')}")
            else:
                print("leer")
    else:
        print(f"Listendaten ausreichend ({len(properties)} Objekte)")
    if not properties:
        print("Keine Objekte gefunden")
        sys.exit(1)
    def sort_key(p):
        nr = str(p.get("objnr", "") or p.get("id", "0"))
        try:
            return (0, int(nr))
        except ValueError:
            return (1, nr)
    properties.sort(key=sort_key)
    print(f"Schreibe {args.out} ...")
    write_xlsx(properties, Path(args.out))
    print("Fertig")


if __name__ == "__main__":
    main()
